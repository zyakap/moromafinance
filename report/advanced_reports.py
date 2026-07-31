import csv
import datetime
from collections import OrderedDict
from decimal import Decimal

from django.shortcuts import render
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDay, TruncMonth, TruncWeek, TruncYear
from django.http import Http404, HttpResponse
from django.utils import timezone

from loan.models import Loan, Payment, Statement, AlescoPayLine
from admin1.models import get_enabled_deduction_categories, deduction_category_label


REPORT_CATEGORIES = [
    {
        'title': 'Loan Portfolio Reports',
        'slug': 'loan-portfolio',
        'reports': [
            ('total-loan-disbursement', 'Total Loan Disbursement'),
            ('outstanding-principal-interest', 'Outstanding Principal and Interest'),
            ('interest-earned-unearned', 'Interest Earned and Unearned'),
            ('active-vs-closed-loans', 'Active Loans vs Closed Loans'),
            ('loan-distribution-employer', 'Loan Distribution by Employer'),
            ('loans-by-referrer', 'Loans by Referrer'),
            ('loan-distribution-region', 'Loan Distribution by Region'),
            ('loan-distribution-product', 'Loan Distribution by Product'),
            ('loans-outstanding-company', 'Loans Outstanding by Department/Company'),
        ],
    },
    {
        'title': 'Loan Disbursement Reports',
        'slug': 'loan-disbursements',
        'reports': [
            ('disbursements-daily', 'Daily Disbursements'),
            ('disbursements-weekly', 'Weekly Disbursements'),
            ('disbursements-monthly', 'Monthly Disbursements'),
            ('disbursements-annually', 'Annual Disbursements'),
            ('disbursement-by-client', 'Disbursement by Client'),
            ('disbursement-by-employer', 'Disbursement by Employer'),
            ('new-vs-repeat-loans', 'New Loans vs Repeat Loans'),
            ('processing-fees', 'Processing Fees'),
        ],
    },
    {
        'title': 'Repayments & Collection Reports',
        'slug': 'repayments-collections',
        'reports': [
            ('expected-this-pay', 'Total Expected This Pay'),
            ('expected-next-pay', 'Total Expected Next Pay'),
            ('expected-by-employer', 'Expected Collections by Employer'),
            ('clients-by-deduction-category', 'Clients by Deduction Category'),
            ('payroll-deduction-alesco', 'Payroll Deduction Alesco Report'),
            ('collections-weekly', 'Weekly Collections'),
            ('collections-fortnightly', 'Fortnightly Collections'),
            ('collections-monthly', 'Monthly Collections'),
            ('collections-total', 'Total Collected Principal and Interest'),
            ('expected-vs-actual-repayments', 'Expected vs Actual Repayments'),
            ('expected-vs-actual-by-client', 'Expected vs Actual by Client'),
        ],
    },
    {
        'title': 'Alesco Payroll Reports',
        'slug': 'alesco-payroll',
        'reports': [
            ('alesco-per-pay-period', 'Payments vs Expected — Per Pay Period'),
            ('alesco-per-employer', 'Payments vs Expected — Per Employer'),
            ('alesco-unmatched-categorized', 'Unmatched Payments — Categorized'),
        ],
    },
    {
        'title': 'Arrears & Delinquency Reports',
        'slug': 'arrears-delinquency',
        'reports': [
            ('arrears-0-30', '0-30 Days Arrears'),
            ('arrears-31-60', '31-60 Days Arrears'),
            ('arrears-61-90', '61-90 Days Arrears'),
            ('arrears-90-plus', '90+ Days Arrears'),
            ('non-performing-loans', 'Non-Performing Loans'),
            ('defaulted-loans-report', 'Defaulted Loans'),
            ('default-interest-report', 'Default Interest'),
            ('missed-payroll-deductions', 'Missed Payroll Deductions'),
        ],
    },
    {
        'title': 'Interest & Income Reports',
        'slug': 'interest-income',
        'reports': [
            ('interest-accrued', 'Interest Accrued'),
            ('interest-collected', 'Interest Collected'),
            ('loan-income-by-product', 'Loan Income by Product'),
            ('loan-income-by-client', 'Loan Income by Client'),
            ('loan-income-by-employer', 'Loan Income by Employer'),
        ],
    },
    {
        'title': 'Client Level Reports',
        'slug': 'client-level',
        'reports': [
            ('client-loan-statement', 'Client Loan Statement'),
            ('loan-amortization-schedule', 'Loan Amortization Schedule'),
            ('client-outstanding-balance', 'Outstanding Balance per Client'),
            ('client-repayment-history', 'Client Repayment History'),
            ('top-borrowers', 'Top Borrowers'),
            ('repeat-borrowers', 'Repeat Borrowers'),
        ],
    },
    {
        'title': 'Recovery & Write-Off Reports',
        'slug': 'recovery-writeoff',
        'reports': [
            ('recovery-collections', 'Recovery Collections'),
            ('write-off-report', 'Write-Off Report'),
            ('recovered-loans-report', 'Recovered Loans'),
        ],
    },
    {
        'title': 'Financial & Accounting Reports',
        'slug': 'financial-accounting',
        'reports': [
            ('loan-disbursement-summary', 'Loan Disbursement Summary'),
            ('repayments-summary', 'Repayments Summary'),
            ('processing-fees-summary', 'Processing Fees Summary'),
            ('outstanding-loan-report', 'Outstanding Loan Report'),
            ('provision-bad-debts', 'Provision for Bad Debts'),
        ],
    },
    {
        'title': 'Management Dashboard Reports',
        'slug': 'management-dashboard',
        'reports': [
            ('business-performance', 'Business Performance Overview'),
            ('portfolio-report', 'Portfolio Report'),
            ('collection-rate', 'Collection Rate'),
            ('default-rate', 'Default Rate'),
            ('profitability-report', 'Profitability Report'),
            ('growth-trends', 'Growth Trends'),
        ],
    },
    {
        'title': 'Processing Fee Collections',
        'slug': 'processing-fee-collections',
        'reports': [
            ('pfc-weekly', 'Weekly Processing Fee Collections'),
            ('pfc-fortnightly', 'Fortnightly Processing Fee Collections'),
            ('pfc-monthly', 'Monthly Processing Fee Collections'),
            ('pfc-annually', 'Annual Processing Fee Collections'),
        ],
    },
]


REPORT_LOOKUP = {
    report_slug: {'title': report_title, 'category': category['title'], 'category_slug': category['slug']}
    for category in REPORT_CATEGORIES
    for report_slug, report_title in category['reports']
}


def _date_window(request):
    today = timezone.localdate()
    start = request.GET.get('startdate')
    end = request.GET.get('enddate')

    try:
        start_date = datetime.datetime.strptime(start, '%Y-%m-%d').date() if start else today.replace(day=1)
        end_date = datetime.datetime.strptime(end, '%Y-%m-%d').date() if end else today
    except ValueError:
        start_date = today.replace(day=1)
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date


def _D(value):
    """Decimal-safe coercion for arithmetic on possibly-None model fields."""
    if value is None:
        return Decimal('0.00')
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _money(value):
    if value is None:
        value = Decimal('0.00')
    return 'K{:,.2f}'.format(value)


def _number(value):
    if value is None:
        return '0'
    return '{:,}'.format(value)


def _percent(value):
    if value is None:
        value = 0
    return '{:.2f}%'.format(value)


def _date(value):
    return value.strftime('%Y-%m-%d') if value else ''


def _blank_aware_filter(qs, field, value):
    """Filter ``field`` by ``value``; 'Not Specified' matches blank/null so the
    no-employer (or no-group) bucket is drillable and the client identifiable."""
    if value == 'Not Specified':
        return qs.filter(Q(**{f'{field}__isnull': True}) | Q(**{field: ''}))
    return qs.filter(**{field: value})


def _text(value):
    return value if value not in (None, '') else 'Not Specified'


def _format(value, kind='text'):
    if kind == 'money':
        return _money(value)
    if kind == 'number':
        return _number(value)
    if kind == 'percent':
        return _percent(value)
    if kind == 'date':
        return _date(value)
    return _text(value)


def _sum(queryset, field):
    return queryset.aggregate(total=Sum(field))['total'] or Decimal('0.00')


def _loan_base():
    return Loan.objects.select_related('owner', 'location').filter(category='FUNDED')


def _payment_base():
    return Payment.objects.select_related('owner', 'loanref').all()


def _statement_base():
    return Statement.objects.select_related('owner', 'loanref').all()


def _loan_summary(loans):
    principal = _sum(loans, 'principal_loan_receivable')
    ordinary_interest = _sum(loans, 'ordinary_interest_receivable')
    default_interest = _sum(loans, 'default_interest_receivable')
    principal_paid = _sum(loans, 'principal_loan_paid')
    interest_paid = _sum(loans, 'interest_paid')
    total_outstanding = _sum(loans, 'total_outstanding')
    total_disbursed = _sum(loans, 'amount')
    total_collected = principal_paid + interest_paid + _sum(loans, 'default_interest_paid')

    return [
        ('Loans', _number(loans.count())),
        ('Disbursed', _money(total_disbursed)),
        ('Principal Outstanding', _money(principal)),
        ('Interest Outstanding', _money(ordinary_interest + default_interest)),
        ('Total Outstanding', _money(total_outstanding)),
        ('Collected', _money(total_collected)),
    ]


def _client_name(item):
    name = '{} {}'.format(item.get('owner__first_name') or '', item.get('owner__last_name') or '').strip()
    return name or _text(item.get('owner__email'))


def _loan_rows(loans):
    fields = [
        ('Ref', 'ref', 'text'),
        ('Client', None, 'name'),
        ('Employer', 'owner__employer', 'text'),
        ('Funding Date', 'funding_date', 'date'),
        ('Principal', 'amount', 'money'),
        ('Principal Outstanding', 'principal_loan_receivable', 'money'),
        ('Interest Outstanding', 'ordinary_interest_receivable', 'money'),
        ('Default Interest', 'default_interest_receivable', 'money'),
        ('Total Outstanding', 'total_outstanding', 'money'),
        ('Arrears', 'total_arrears', 'money'),
        ('Status', 'status', 'text'),
    ]
    value_fields = ['ref', 'owner__first_name', 'owner__last_name', 'owner__email', 'owner__employer',
                    'funding_date', 'amount', 'principal_loan_receivable', 'ordinary_interest_receivable',
                    'default_interest_receivable', 'total_outstanding', 'total_arrears', 'status']
    values = loans.values(*value_fields)[:500]
    rows = []
    for item in values:
        row = []
        for label, key, kind in fields:
            if kind == 'name':
                row.append(_client_name(item))
            else:
                row.append(_format(item.get(key), kind))
        rows.append(row)
    return [label for label, _, _ in fields], rows


def _expected_repayment_rows(loans):
    """Expected repayments for a pay window — one row per loan due, with the
    expected repayment amount, so the totals reconcile against the pay run."""
    loans = loans.select_related('owner').order_by('owner__employer', 'owner__last_name', 'owner__first_name')
    fields = ['ref', 'owner__first_name', 'owner__last_name', 'owner__email', 'owner__employer',
              'next_payment_date', 'repayment_amount', 'total_outstanding', 'total_arrears', 'status']
    rows = []
    for item in loans.values(*fields)[:1000]:
        rows.append([
            _text(item.get('ref')),
            _client_name(item),
            _text(item.get('owner__employer')),
            _date(item.get('next_payment_date')),
            _money(item.get('repayment_amount')),
            _money(item.get('total_outstanding')),
            _money(item.get('total_arrears')),
            _text(item.get('status')),
        ])
    columns = ['Ref', 'Client', 'Employer', 'Next Payment', 'Expected Repayment', 'Total Outstanding', 'Arrears', 'Status']
    return columns, rows


def _expected_by_employer_rows(loans):
    """Expected repayments grouped by employer — how much to expect from each
    employer this pay (number of loans due and total expected)."""
    grouped = loans.values('owner__employer').annotate(
        count=Count('id'),
        expected=Sum('repayment_amount'),
        outstanding=Sum('total_outstanding'),
        arrears=Sum('total_arrears'),
    ).order_by('-expected')
    rows = []
    for item in grouped:
        rows.append([
            _text(item.get('owner__employer')),
            _number(item.get('count')),
            _money(item.get('expected')),
            _money(item.get('outstanding')),
            _money(item.get('arrears')),
        ])
    return ['Employer', 'Loans Due', 'Expected This Pay', 'Total Outstanding', 'Arrears'], rows


def _grouped_loan_rows(loans, group_field, group_label):
    rows = []
    grouped = loans.values(group_field).annotate(
        count=Count('id'),
        disbursed=Sum('amount'),
        principal=Sum('principal_loan_receivable'),
        interest=Sum('ordinary_interest_receivable'),
        default_interest=Sum('default_interest_receivable'),
        outstanding=Sum('total_outstanding'),
        arrears=Sum('total_arrears'),
    ).order_by(group_field)

    for item in grouped:
        rows.append([
            _text(item.get(group_field)),
            _number(item.get('count')),
            _money(item.get('disbursed')),
            _money(item.get('principal')),
            _money(item.get('interest')),
            _money(item.get('default_interest')),
            _money(item.get('outstanding')),
            _money(item.get('arrears')),
        ])

    return [group_label, 'Loans', 'Disbursed', 'Principal Outstanding', 'Interest Outstanding', 'Default Interest', 'Total Outstanding', 'Arrears'], rows


def _periodic_loan_rows(loans, trunc):
    truncators = {
        'day': TruncDay,
        'week': TruncWeek,
        'month': TruncMonth,
        'year': TruncYear,
    }
    grouped = loans.annotate(period=truncators[trunc]('funding_date')).values('period').annotate(
        count=Count('id'),
        disbursed=Sum('amount'),
        processing_fees=Sum('processing_fee'),
        principal=Sum('principal_loan_receivable'),
        interest=Sum('ordinary_interest_receivable'),
    ).order_by('period')
    rows = [[_date(item['period']), _number(item['count']), _money(item['disbursed']), _money(item['processing_fees']), _money(item['principal']), _money(item['interest'])] for item in grouped]
    return ['Period', 'Loans', 'Disbursed', 'Processing Fees', 'Principal Outstanding', 'Interest Outstanding'], rows


def _period_bounds(period_str, trunc):
    """Start/end funding dates for one period bucket (matching _periodic_loan_rows'
    TruncDay/Week/Month/Year). Used to drill a period summary row into its loans."""
    import calendar
    d = datetime.datetime.strptime(str(period_str)[:10], '%Y-%m-%d').date()
    if trunc == 'week':
        return d, d + datetime.timedelta(days=6)
    if trunc == 'month':
        return d.replace(day=1), d.replace(day=calendar.monthrange(d.year, d.month)[1])
    if trunc == 'year':
        return d.replace(month=1, day=1), d.replace(month=12, day=31)
    return d, d  # day


# Periodic disbursement reports (slug -> period granularity) whose period rows
# drill down to the individual loans disbursed in that period.
_DISBURSEMENT_PERIODIC = {
    'disbursements-daily': 'day',
    'disbursements-weekly': 'week',
    'disbursements-monthly': 'month',
    'disbursements-annually': 'year',
}

# Grouped reports whose first-column group value drills down to the individual
# loans in that group. slug -> (loan filter field, only-with-outstanding-balance).
_GROUP_DRILL_FIELDS = {
    'loans-by-referrer': ('owner__referred_by__name', False),
    'active-vs-closed-loans': ('funded_category', False),
    'loan-distribution-region': ('owner__residential_province', False),
    'loan-distribution-product': ('loan_type', False),
    'loans-outstanding-company': ('owner__employer', True),
    'disbursement-by-client': ('owner__email', False),
    'loan-income-by-client': ('owner__email', False),
    'client-outstanding-balance': ('owner__email', True),
}


def _processing_fee_detail_rows(loans, trunc):
    """Return client-level processing fee rows grouped by period, showing date, client, department, fee."""
    truncators = {
        'week': TruncWeek,
        'fortnight': TruncWeek,
        'month': TruncMonth,
        'year': TruncYear,
    }
    fee_loans = loans.filter(processing_fee__gt=0).select_related('owner').order_by('funding_date')

    # Build period label -> list of detail rows
    from collections import OrderedDict
    period_rows = []
    total_fees = Decimal('0.00')

    for loan in fee_loans[:500]:
        client_name = f'{loan.owner.first_name} {loan.owner.last_name}'.strip() if loan.owner else 'Unknown'
        department = _text(loan.owner.employer if loan.owner else None)
        fee = loan.processing_fee or Decimal('0.00')
        total_fees += fee
        period_rows.append([
            _text(loan.ref),
            _date(loan.funding_date),
            client_name,
            department,
            _money(loan.amount),
            _money(fee),
        ])

    return ['Ref', 'Date', 'Client Name', 'Department', 'Loan Amount', 'Fee Amount'], period_rows, total_fees


def _periodic_payment_rows(payments, trunc):
    truncators = {
        'day': TruncDay,
        'week': TruncWeek,
        'fortnight': TruncWeek,
        'month': TruncMonth,
        'year': TruncYear,
    }
    grouped = payments.annotate(period=truncators[trunc]('date')).values('period').annotate(
        count=Count('id'),
        collected=Sum('amount'),
    ).order_by('period')
    rows = [[_date(item['period']), _number(item['count']), _money(item['collected'])] for item in grouped]
    return ['Period', 'Payments', 'Collected'], rows


def _collections_detail_rows(payments):
    """Full per-payment detail for the period — one row per repayment collected,
    so the report can be reconciled line-by-line against the Alesco pay listing."""
    payments = payments.select_related('owner', 'loanref').order_by(
        'date', 'owner__employer', 'owner__last_name', 'owner__first_name')
    rows = []
    total = Decimal('0.00')
    for p in payments[:2000]:
        client = f'{p.owner.first_name} {p.owner.last_name}'.strip() if p.owner else 'Unknown'
        amount = p.amount or Decimal('0.00')
        total += amount
        rows.append([
            _date(p.date),
            _text(p.loanref.ref if p.loanref else None),
            client,
            _text(p.owner.employer if p.owner else None),
            _text(p.statement),
            _text(p.mode),
            _money(amount),
        ])
    columns = ['Date', 'Loan Ref', 'Client Name', 'Employer', 'Description', 'Mode', 'Amount']
    return columns, rows, total


def _statement_collection_rows(statements):
    grouped = statements.values('loanref__owner__employer').annotate(
        count=Count('id'),
        principal=Sum('principal_collected'),
        interest=Sum('interest_collected'),
        default_interest=Sum('default_interest_collected'),
        credit=Sum('credit'),
    ).order_by('loanref__owner__employer')
    rows = []
    for item in grouped:
        total = (item.get('principal') or 0) + (item.get('interest') or 0) + (item.get('default_interest') or 0)
        rows.append([
            _text(item.get('loanref__owner__employer')),
            _number(item.get('count')),
            _money(item.get('principal')),
            _money(item.get('interest')),
            _money(item.get('default_interest')),
            _money(item.get('credit') or total),
        ])
    return ['Employer', 'Transactions', 'Principal', 'Interest', 'Default Interest', 'Total Collected'], rows


def _scheduled_expected_in_window(loan, start_date, end_date):
    """(expected amount, count of due fortnights) whose SCHEDULED due date falls
    in the window — read from the immutable schedule, independent of whether the
    fortnight has been paid. This is why Expected does not clear once paid."""
    from loan import schedule as _sched
    dates = _sched.build_schedule(loan)
    amounts = _sched.expected_amounts(loan)
    total = Decimal('0.00')
    count = 0
    for i, d in enumerate(dates):
        if start_date <= d <= end_date:
            total += amounts[i] if i < len(amounts) else Decimal('0')
            count += 1
    return total, count


def _expected_vs_actual_rows(start_date, end_date, payroll_only=False, employer=None, by_client=False):
    """Expected (from the immutable schedule) vs Actual (payments received) for the
    window. Expected persists after payment, so variance stays meaningful. When
    ``employer`` is given, return the client-level breakdown for that employer;
    with ``by_client`` return one row per loan account across ALL employers
    (who paid and who did not, without drilling)."""
    funded = (_loan_base().filter(category='FUNDED')
              .exclude(funded_category__in=['COMPLETED', 'ARCHIVED']).select_related('owner'))
    payments = _payment_base().filter(date__gte=start_date, date__lte=end_date)
    if payroll_only:
        payments = payments.filter(mode='PAYROLL DEDUCTION')
    if employer is not None:
        funded = _blank_aware_filter(funded, 'owner__employer', employer)
        payments = _blank_aware_filter(payments, 'owner__employer', employer)

    if employer is not None or by_client:
        # Client-level: one row per loan (drill-down or the flat all-clients report).
        actual_by_loan = {p['loanref']: (p['a'] or Decimal('0.00'))
                          for p in payments.values('loanref').annotate(a=Sum('amount'))}
        rows = []
        for loan in funded:
            exp, _cnt = _scheduled_expected_in_window(loan, start_date, end_date)
            act = actual_by_loan.get(loan.id, Decimal('0.00'))
            if exp == 0 and act == 0:
                continue
            variance = act - exp
            rate = (act / exp * 100) if exp else 0
            client = f'{loan.owner.first_name} {loan.owner.last_name}'.strip() if loan.owner else 'Unknown'
            # Advance Payments (+) / Arrears (-): the loan's CURRENT standing —
            # not window-limited — so a defaulting client's existing arrears (or
            # a good client's prepaid advance credit) is visible at a glance.
            adv_arrears = _D(loan.advance_balance) - _D(loan.total_arrears)
            if by_client and employer is None:
                emp = _text(loan.owner.employer if loan.owner else None)
                rows.append([_text(loan.ref), client, emp, _money(exp), _money(act),
                             _money(variance), _money(adv_arrears), _percent(rate)])
            else:
                rows.append([_text(loan.ref), client, _money(exp), _money(act), _money(variance),
                             _money(adv_arrears), _percent(rate)])
        if by_client and employer is None:
            rows.sort(key=lambda r: (r[1] or '').upper())
            return (['Ref', 'Client', 'Employer', 'Expected', 'Actual', 'Variance',
                     'Advance / (Arrears)', 'Collection Rate'], rows)
        return (['Ref', 'Client', 'Expected', 'Actual', 'Variance',
                 'Advance / (Arrears)', 'Collection Rate'], rows)

    # Grouped by employer.
    employers = OrderedDict()
    for loan in funded:
        emp = _text(loan.owner.employer if loan.owner else None)
        exp, cnt = _scheduled_expected_in_window(loan, start_date, end_date)
        e = employers.setdefault(emp, {'expected': Decimal('0.00'), 'actual': Decimal('0.00'), 'loans': 0})
        e['expected'] += exp
        if cnt > 0:
            e['loans'] += 1
    for item in payments.values('owner__employer').annotate(actual=Sum('amount')):
        emp = _text(item.get('owner__employer'))
        e = employers.setdefault(emp, {'expected': Decimal('0.00'), 'actual': Decimal('0.00'), 'loans': 0})
        e['actual'] = item.get('actual') or Decimal('0.00')

    # Current advance/arrears standing per employer (not window-limited),
    # so directors can see at a glance which employer groups carry arrears.
    adv_arrears_by_emp = OrderedDict()
    for loan in funded:
        emp = _text(loan.owner.employer if loan.owner else None)
        adv_arrears_by_emp[emp] = adv_arrears_by_emp.get(emp, Decimal('0.00')) + (
            _D(loan.advance_balance) - _D(loan.total_arrears))

    rows = []
    for emp, t in employers.items():
        if t['expected'] == 0 and t['actual'] == 0:
            continue  # no activity in the window — omit
        variance = t['actual'] - t['expected']
        rate = (t['actual'] / t['expected'] * 100) if t['expected'] else 0
        adv_arrears = adv_arrears_by_emp.get(emp, Decimal('0.00'))
        rows.append([emp, _number(t['loans']), _money(t['expected']), _money(t['actual']), _money(variance),
                     _money(adv_arrears), _percent(rate)])
    rows.sort(key=lambda r: r[0])
    return (['Employer', 'Loans Due', 'Expected', 'Actual', 'Variance',
             'Advance / (Arrears)', 'Collection Rate'], rows)


def _loans_by_referrer_rows(loans):
    """Portfolio grouped by referrer — who refers GOOD clients: per referrer the
    clients/loans referred, money disbursed and outstanding, arrears, defaulted
    loans and the share of their book that is in good standing."""
    buckets = OrderedDict()
    for loan in loans.select_related('owner', 'owner__referred_by'):
        ref = loan.owner.referred_by if (loan.owner and loan.owner.referred_by_id) else None
        key = ref.name if ref else '— No referrer —'
        b = buckets.setdefault(key, {'clients': set(), 'loans': 0, 'disbursed': Decimal('0.00'),
                                     'outstanding': Decimal('0.00'), 'arrears': Decimal('0.00'),
                                     'defaulted': 0, 'good': 0})
        if loan.owner_id:
            b['clients'].add(loan.owner_id)
        b['loans'] += 1
        b['disbursed'] += (loan.amount or 0)
        b['outstanding'] += (loan.total_outstanding or 0)
        b['arrears'] += (loan.total_arrears or 0)
        if loan.status == 'DEFAULTED' or (loan.total_arrears or 0) > 0:
            b['defaulted'] += 1
        else:
            b['good'] += 1
    rows = []
    for name, b in buckets.items():
        good_pct = (Decimal(b['good']) / Decimal(b['loans']) * 100) if b['loans'] else Decimal('0')
        rows.append([name, _number(len(b['clients'])), _number(b['loans']), _money(b['disbursed']),
                     _money(b['outstanding']), _money(b['arrears']), _number(b['defaulted']),
                     _percent(good_pct)])
    # best referrers (highest good-loan share, then volume) first; No-referrer bucket last
    rows.sort(key=lambda r: (r[0] == '— No referrer —',
                             -float(str(r[7]).replace('%', '') or 0), -int(r[2] or 0)))
    return (['Referrer', 'Clients', 'Loans', 'Disbursed', 'Outstanding',
             'Arrears', 'Loans in Default/Arrears', 'Good Loans %'], rows)


def _alesco_line_rows(start_date, end_date, employer=None, group_by_employer=False):
    """Confirmed Alesco payroll deductions in the pay-period window (filtered on
    the run's Period End date). Returns payments vs expected repayment + variance.

    * group_by_employer=False → one row per client (detail listing).
    * group_by_employer=True  → one aggregated row per employer.
    """
    lines = (AlescoPayLine.objects
             .select_related('run', 'owner', 'loanref')
             .filter(status='CONFIRMED',
                     run__period_end__gte=start_date,
                     run__period_end__lte=end_date))
    if employer:
        lines = _blank_aware_filter(lines, 'employer_name', employer)

    if group_by_employer:
        buckets = OrderedDict()
        for ln in lines:
            emp = _text(ln.employer_name)
            b = buckets.setdefault(emp, {'clients': set(), 'paid': Decimal('0.00'),
                                         'expected': Decimal('0.00')})
            if ln.owner_id:
                b['clients'].add(ln.owner_id)
            b['paid'] += (ln.this_period or 0)
            b['expected'] += (ln.expected_repayment or 0)
        rows = []
        for emp, b in buckets.items():
            variance = b['paid'] - b['expected']
            rows.append([emp, _number(len(b['clients'])), _money(b['expected']),
                         _money(b['paid']), _money(variance)])
        return ['Employer', 'Clients', 'Expected', 'Paid (This Period)', 'Variance'], rows

    rows = []
    for ln in lines.order_by('run__period_end', 'employer_name', 'report_name'):
        client = (f'{ln.owner.first_name} {ln.owner.last_name}'.strip()
                  if ln.owner else ln.report_name)
        rows.append([
            _date(ln.run.period_end),
            ln.employee_file_number or '',
            _text(client),
            _text(ln.employer_name),
            (ln.loanref.ref if ln.loanref else ''),
            _money(ln.this_period),
            _money(ln.expected_repayment),
            _money(ln.variance),
        ])
    columns = ['Pay Period End', 'Employee File', 'Client', 'Employer', 'LoanRef',
               'Paid (This Period)', 'Expected', 'Variance']
    return columns, rows


def _clients_by_deduction_category_rows():
    """How many clients (and their loan book) fall under each payroll/deduction
    category — so admin can see e.g. how many ALESCO vs Private Sector clients."""
    from accounts.models import UserProfile
    rows = []
    for value, label in get_enabled_deduction_categories():
        clients = UserProfile.objects.filter(category='CLIENT', deduction_category=value).count()
        book = (_loan_base().filter(owner__deduction_category=value, category='FUNDED')
                .exclude(funded_category__in=['COMPLETED', 'ARCHIVED']))
        rows.append([label, _number(clients), _number(book.count()),
                     _money(_sum(book, 'total_outstanding')), _money(_sum(book, 'total_arrears'))])
    unassigned = UserProfile.objects.filter(category='CLIENT').filter(
        Q(deduction_category__isnull=True) | Q(deduction_category='')).count()
    if unassigned:
        book = (_loan_base().filter(category='FUNDED').filter(
                Q(owner__deduction_category__isnull=True) | Q(owner__deduction_category=''))
                .exclude(funded_category__in=['COMPLETED', 'ARCHIVED']))
        rows.append(['Unassigned', _number(unassigned), _number(book.count()),
                     _money(_sum(book, 'total_outstanding')), _money(_sum(book, 'total_arrears'))])
    return ['Payroll Category', 'Clients', 'Active Loans', 'Outstanding', 'Arrears'], rows


def _business_performance_rows(start_date, end_date):
    """Consolidated lending KPIs: portfolio health (as of now) + business
    performance (for the selected period). Returns a two-column metric table
    with section-header rows."""
    loans = _loan_base()
    funded = loans.filter(category='FUNDED')
    active = funded.filter(funded_category='ACTIVE')
    book = funded.exclude(funded_category__in=['COMPLETED', 'ARCHIVED'])
    npl = funded.filter(Q(status='DEFAULTED') | Q(funded_category__in=['RECOVERY', 'BAD', 'WOFF']))
    in_arrears = book.filter(total_arrears__gt=0)
    par30 = book.filter(days_in_default__gt=30)

    gross = _sum(book, 'total_outstanding')
    par30_out = _sum(par30, 'total_outstanding')
    funded_count = funded.count()

    period_loans = funded.filter(funding_date__gte=start_date, funding_date__lte=end_date)
    new_loans = period_loans.filter(classification='NEW')
    repeat_loans = period_loans.exclude(classification='NEW')

    # Collections in the period, split by component (from PAYMENT statements).
    period_pmts = _statement_base().filter(date__gte=start_date, date__lte=end_date, type='PAYMENT')
    principal_col = _sum(period_pmts, 'principal_collected')
    interest_col = _sum(period_pmts, 'interest_collected')
    default_int_col = _sum(period_pmts, 'default_interest_collected')
    actual = _sum(_payment_base().filter(date__gte=start_date, date__lte=end_date), 'amount')
    expected = _sum(funded.filter(next_payment_date__gte=start_date, next_payment_date__lte=end_date), 'repayment_amount')
    fee_income = _sum(period_loans, 'processing_fee')

    avg_loan = (_sum(funded, 'amount') / funded_count) if funded_count else 0

    def pct(n, d):
        return (n / d * 100) if d else 0

    metrics = [
        ('— PORTFOLIO HEALTH (as of today) —', '', 'header'),
        ('Active Loans', active.count(), 'number'),
        ('Gross Loan Portfolio (outstanding)', gross, 'money'),
        ('Principal Outstanding', _sum(book, 'principal_loan_receivable'), 'money'),
        ('Interest Outstanding', _sum(book, 'ordinary_interest_receivable') + _sum(book, 'default_interest_receivable'), 'money'),
        ('Loans in Arrears', in_arrears.count(), 'number'),
        ('Total Arrears', _sum(book, 'total_arrears'), 'money'),
        ('Portfolio at Risk (PAR>30 days)', pct(par30_out, gross), 'percent'),
        ('Default / NPL Rate', pct(npl.count(), funded_count), 'percent'),
        ('Average Loan Size', avg_loan, 'money'),
        ('— BUSINESS PERFORMANCE (selected period) —', '', 'header'),
        ('Loans Disbursed', period_loans.count(), 'number'),
        ('Amount Disbursed', _sum(period_loans, 'amount'), 'money'),
        ('New vs Repeat Disbursement', f"{_money(_sum(new_loans,'amount'))} / {_money(_sum(repeat_loans,'amount'))}", 'text'),
        ('Expected Collections', expected, 'money'),
        ('Actual Collections', actual, 'money'),
        ('Collection Rate', pct(actual, expected), 'percent'),
        ('Principal Collected', principal_col, 'money'),
        ('Interest Collected', interest_col, 'money'),
        ('Default Interest Collected', default_int_col, 'money'),
        ('Processing Fee Income', fee_income, 'money'),
        ('Total Income (interest + default int + fees)', interest_col + default_int_col + fee_income, 'money'),
    ]
    rows = [[label, ('' if kind == 'header' else _format(value, kind))] for label, value, kind in metrics]
    return ['Metric', 'Value'], rows


def _dashboard_rows(slug, start_date, end_date):
    loans = _loan_base()
    period_loans = loans.filter(funding_date__gte=start_date, funding_date__lte=end_date)
    period_payments = _payment_base().filter(date__gte=start_date, date__lte=end_date)
    active = loans.filter(funded_category='ACTIVE')
    closed = loans.filter(funded_category__in=['COMPLETED', 'ARCHIVED'])
    defaulted = loans.filter(Q(status='DEFAULTED') | Q(funded_category__in=['RECOVERY', 'BAD', 'WOFF']))
    expected = _sum(loans.filter(next_payment_date__gte=start_date, next_payment_date__lte=end_date), 'repayment_amount')
    actual = _sum(period_payments, 'amount')
    income = _sum(period_loans, 'processing_fee') + _sum(period_payments, 'amount')

    data = {
        'portfolio-report': [
            ('Active Loans', active.count(), 'number'),
            ('Closed Loans', closed.count(), 'number'),
            ('Portfolio Outstanding', _sum(loans, 'total_outstanding'), 'money'),
            ('Period Disbursement', _sum(period_loans, 'amount'), 'money'),
        ],
        'collection-rate': [
            ('Expected', expected, 'money'),
            ('Actual', actual, 'money'),
            ('Collection Rate', (actual / expected * 100) if expected else 0, 'percent'),
        ],
        'default-rate': [
            ('Funded Loans', loans.count(), 'number'),
            ('Defaulted/NPL Loans', defaulted.count(), 'number'),
            ('Default Rate', (defaulted.count() / loans.count() * 100) if loans.count() else 0, 'percent'),
            ('Arrears', _sum(defaulted, 'total_arrears'), 'money'),
        ],
        'profitability-report': [
            ('Processing Fees', _sum(period_loans, 'processing_fee'), 'money'),
            ('Interest Collected', _sum(_statement_base().filter(date__gte=start_date, date__lte=end_date, type='PAYMENT'), 'interest_collected'), 'money'),
            ('Total Collections', _sum(period_payments, 'amount'), 'money'),
            ('Interest Receivable', _sum(loans, 'ordinary_interest_receivable') + _sum(loans, 'default_interest_receivable'), 'money'),
        ],
        'growth-trends': [
            ('New Loans', period_loans.filter(classification='NEW').count(), 'number'),
            ('Repeat Loans', period_loans.exclude(classification='NEW').count(), 'number'),
            ('New Disbursement', _sum(period_loans.filter(classification='NEW'), 'amount'), 'money'),
            ('Repeat Disbursement', _sum(period_loans.exclude(classification='NEW'), 'amount'), 'money'),
        ],
    }
    rows = [[label, _format(value, kind)] for label, value, kind in data[slug]]
    return ['Metric', 'Value'], rows


def _period_key(d, trunc, anchor):
    """Bucket a date into its period start for weekly/fortnightly/monthly/annual
    grouping. Fortnights are 14-day bins anchored at ``anchor`` (the report start)."""
    if not d:
        return None
    if trunc == 'week':
        return d - datetime.timedelta(days=d.weekday())
    if trunc == 'fortnight':
        anchor = anchor or d
        delta = (d - anchor).days
        return anchor + datetime.timedelta(days=(delta // 14) * 14)
    if trunc == 'month':
        return d.replace(day=1)
    if trunc == 'year':
        return d.replace(month=1, day=1)
    return d


def _periodic_collection_rows(statements, trunc, anchor):
    """Collections aggregated per period (week/fortnight/month), split into
    principal, interest and default interest, with the total collected."""
    buckets = OrderedDict()
    qs = statements.filter(type='PAYMENT')
    for s in qs:
        k = _period_key(s.date, trunc, anchor)
        if k is None:
            continue
        b = buckets.setdefault(k, {'n': 0, 'p': Decimal('0'), 'i': Decimal('0'),
                                   'd': Decimal('0'), 't': Decimal('0')})
        b['n'] += 1
        b['p'] += s.principal_collected or 0
        b['i'] += s.interest_collected or 0
        b['d'] += s.default_interest_collected or 0
        b['t'] += s.debit or 0
    rows = [[(k.isoformat() if k else '—'), _number(v['n']), _money(v['p']), _money(v['i']), _money(v['d']), _money(v['t'])]
            for k, v in sorted(buckets.items(), key=lambda kv: (kv[0] is None, kv[0]))]
    return ['Period', 'Payments', 'Principal', 'Interest', 'Default Interest', 'Total Collected'], rows


def _periodic_fee_rows(loans, trunc, anchor):
    """Processing fees aggregated per period (week/fortnight/month/year)."""
    buckets = OrderedDict()
    total = Decimal('0.00')
    for l in loans.filter(processing_fee__gt=0):
        k = _period_key(l.funding_date, trunc, anchor)
        if k is None:
            continue
        b = buckets.setdefault(k, {'n': 0, 'disb': Decimal('0'), 'fee': Decimal('0')})
        b['n'] += 1
        b['disb'] += l.amount or 0
        b['fee'] += l.processing_fee or 0
        total += l.processing_fee or 0
    rows = [[_date(k), _number(v['n']), _money(v['disb']), _money(v['fee'])]
            for k, v in sorted(buckets.items(), key=lambda kv: (kv[0] is None, kv[0]))]
    return ['Period', 'Loans', 'Disbursed', 'Processing Fees'], rows, total


def _statement_collection_by_client_rows(statements):
    """Per-client (per-loan) collections summary — one row per loan showing the
    principal/interest/default interest collected and total. A genuine client
    loan statement summary (as opposed to an employer roll-up)."""
    grouped = statements.filter(type='PAYMENT').values(
        'loanref__ref', 'loanref__owner__first_name', 'loanref__owner__last_name',
        'loanref__owner__employer',
    ).annotate(
        count=Count('id'),
        principal=Sum('principal_collected'),
        interest=Sum('interest_collected'),
        default_interest=Sum('default_interest_collected'),
        total=Sum('debit'),
    ).order_by('loanref__owner__last_name', 'loanref__owner__first_name')
    rows = []
    for item in grouped:
        name = '{} {}'.format(item.get('loanref__owner__first_name') or '',
                              item.get('loanref__owner__last_name') or '').strip()
        rows.append([
            _text(item.get('loanref__ref')),
            _text(name),
            _text(item.get('loanref__owner__employer')),
            _number(item.get('count')),
            _money(item.get('principal')),
            _money(item.get('interest')),
            _money(item.get('default_interest')),
            _money(item.get('total')),
        ])
    columns = ['Loan Ref', 'Client', 'Employer', 'Payments', 'Principal', 'Interest',
               'Default Interest', 'Total Collected']
    return columns, rows


def _interest_earned_unearned_rows(loans, group_field='loan_type', label='Product'):
    """Earned interest (already collected) vs unearned interest (still
    outstanding/receivable), grouped by product."""
    grouped = loans.values(group_field).annotate(
        count=Count('id'),
        earned=Sum('interest_paid'),
        unearned=Sum('ordinary_interest_receivable'),
        def_earned=Sum('default_interest_paid'),
        def_unearned=Sum('default_interest_receivable'),
    ).order_by(group_field)
    rows = []
    for item in grouped:
        rows.append([
            _text(item.get(group_field)),
            _number(item.get('count')),
            _money(item.get('earned')),
            _money(item.get('unearned')),
            _money(item.get('def_earned')),
            _money(item.get('def_unearned')),
        ])
    return [label, 'Loans', 'Interest Earned', 'Interest Unearned',
            'Default Int Earned', 'Default Int Unearned'], rows


def _interest_accrued_rows(loans, group_field='loan_type', label='Product'):
    """Total interest accrued (charged over the loan life) per product, split into
    collected and still outstanding."""
    grouped = loans.values(group_field).annotate(
        count=Count('id'),
        charged=Sum('interest'),
        collected=Sum('interest_paid'),
        outstanding=Sum('ordinary_interest_receivable'),
    ).order_by(group_field)
    rows = []
    for item in grouped:
        rows.append([
            _text(item.get(group_field)),
            _number(item.get('count')),
            _money(item.get('charged')),
            _money(item.get('collected')),
            _money(item.get('outstanding')),
        ])
    return [label, 'Loans', 'Interest Charged', 'Interest Collected', 'Interest Outstanding'], rows


def _income_by_group_rows(loans, group_field, label):
    """Income lens: interest collected + processing fees per group, with the
    combined income total."""
    grouped = loans.values(group_field).annotate(
        count=Count('id'),
        interest_collected=Sum('interest_paid'),
        default_collected=Sum('default_interest_paid'),
        fees=Sum('processing_fee'),
    ).order_by(group_field)
    rows = []
    for item in grouped:
        ic = item.get('interest_collected') or 0
        dc = item.get('default_collected') or 0
        fees = item.get('fees') or 0
        rows.append([
            _text(item.get(group_field)),
            _number(item.get('count')),
            _money(ic),
            _money(dc),
            _money(fees),
            _money(ic + dc + fees),
        ])
    return [label, 'Loans', 'Interest Collected', 'Default Int Collected',
            'Processing Fees', 'Total Income'], rows


def _report_csv_response(report, slug, start_date, end_date, columns, rows):
    filename = '{}_{}_to_{}.csv'.format(slug, start_date, end_date)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)

    writer = csv.writer(response)
    writer.writerow([report['title']])
    writer.writerow(['Report Category', report['category']])
    writer.writerow(['Start Date', start_date])
    writer.writerow(['End Date', end_date])
    writer.writerow([])
    writer.writerow(columns)
    writer.writerows(rows)
    totals = _compute_totals(columns, rows)
    if totals:
        writer.writerow(totals)

    return response


def _numeric(cell):
    """Parse a formatted cell (e.g. 'K1,234.56', '1,234', '12.5%') to a number,
    or None if it isn't numeric."""
    if cell is None:
        return None
    s = str(cell).strip().replace('K', '').replace(',', '').replace('%', '').strip()
    if s in ('', '-', '—'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _compute_totals(columns, rows):
    """Return a totals row summing the numeric columns (money/number), or None.
    The first column is labelled TOTAL; percentage columns are left blank."""
    if not rows:
        return None
    ncols = len(columns)
    sums = [0.0] * ncols
    has = [False] * ncols
    for row in rows:
        for i in range(min(ncols, len(row))):
            if '%' in str(row[i]):
                continue
            v = _numeric(row[i])
            if v is not None:
                sums[i] += v
                has[i] = True
    if not any(has):
        return None
    totals = []
    for i in range(ncols):
        if i == 0:
            totals.append('TOTAL')
        elif has[i]:
            money = any('K' in str(r[i]) for r in rows if i < len(r))
            totals.append(('K{:,.2f}'.format(sums[i])) if money else ('{:,.0f}'.format(sums[i]) if sums[i] == int(sums[i]) else '{:,.2f}'.format(sums[i])))
        else:
            totals.append('')
    return totals


def _report_xlsx_response(report, slug, start_date, end_date, columns, rows):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'
    ws.append([report['title']])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Category', report['category']])
    ws.append(['Period', '{} to {}'.format(start_date, end_date)])
    ws.append([])
    ws.append(columns)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([(_numeric(c) if _numeric(c) is not None else ('' if c in (None, '—') else str(c))) for c in row])
    totals = _compute_totals(columns, rows)
    if totals:
        ws.append([(_numeric(c) if _numeric(c) is not None else c) for c in totals])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = max(12, min(40, len(str(col)) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="{}_{}_to_{}.xlsx"'.format(slug, start_date, end_date)
    return response


def _report_pdf_response(report, slug, start_date, end_date, columns, rows):
    import subprocess
    from django.utils.html import escape

    head = ''.join('<th>{}</th>'.format(escape(c)) for c in columns)
    body = ''
    for row in rows:
        body += '<tr>' + ''.join('<td>{}</td>'.format(escape(str(c))) for c in row) + '</tr>'
    totals = _compute_totals(columns, rows)
    foot = ''
    if totals:
        foot = '<tr class="total">' + ''.join('<td>{}</td>'.format(escape(str(c))) for c in totals) + '</tr>'
    html = """<html><head><meta charset="utf-8"><style>
      body{{font-family:Arial,Helvetica,sans-serif;font-size:10px;color:#222;}}
      h2{{margin:0 0 2px;}} .meta{{color:#666;font-size:10px;margin-bottom:10px;}}
      table{{width:100%;border-collapse:collapse;}} th,td{{border:1px solid #ccc;padding:4px 6px;text-align:left;}}
      th{{background:#344767;color:#fff;}} tr:nth-child(even) td{{background:#f6f7f9;}}
      tr.total td{{font-weight:bold;background:#eef1f6;}}
    </style></head><body>
      <h2>{title}</h2>
      <div class="meta">{cat} &nbsp;|&nbsp; {start} to {end}</div>
      <table><thead><tr>{head}</tr></thead><tbody>{body}{foot}</tbody></table>
    </body></html>""".format(title=escape(report['title']), cat=escape(report['category']),
                              start=start_date, end=end_date, head=head, body=body, foot=foot)
    proc = subprocess.Popen(
        ['wkhtmltopdf', '-q', '--page-size', 'A4', '--orientation', 'Landscape',
         '--margin-top', '10mm', '--margin-bottom', '10mm', '-', '-'],
        stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    pdf, _err = proc.communicate(html.encode('utf-8'))
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="{}_{}_to_{}.pdf"'.format(slug, start_date, end_date)
    return response


def _maybe_export(request, report, slug, start_date, end_date, columns, rows):
    """Return a download response if ?download=csv|xlsx|pdf, else None."""
    fmt = (request.GET.get('download') or '').lower()
    if fmt == 'csv':
        return _report_csv_response(report, slug, start_date, end_date, columns, rows)
    if fmt in ('xlsx', 'xls', 'excel'):
        return _report_xlsx_response(report, slug, start_date, end_date, columns, rows)
    if fmt == 'pdf':
        return _report_pdf_response(report, slug, start_date, end_date, columns, rows)
    return None


from accounts.functions import staff_or_admin_check


@staff_or_admin_check
def report_catalog(request):
    from admin1.models import referral_program_enabled
    categories = REPORT_CATEGORIES
    if not referral_program_enabled():
        categories = [dict(c, reports=[r for r in c['reports'] if r[0] != 'loans-by-referrer'])
                      for c in REPORT_CATEGORIES]
    from dcc.functions import dcc_enabled
    return render(request, 'report_catalog.html', {
        'nav': 'report_catalog',
        'categories': categories,
        'report_lookup': REPORT_LOOKUP,
        'dcc_enabled': dcc_enabled(),
    })


def _ref_col(columns):
    """Index of the loan-reference column (for linkifying), or None."""
    for name in ('Ref', 'Loan Ref', 'LoanRef'):
        if name in columns:
            return columns.index(name)
    return None


@staff_or_admin_check
def report_detail(request, slug):
    if slug not in REPORT_LOOKUP:
        raise Http404('Report not found')

    start_date, end_date = _date_window(request)
    report = REPORT_LOOKUP[slug]
    loans = _loan_base()

    # Payroll / deduction-category filter: pull reports for only ALESCO, only
    # Private Sector, etc. Applies to every queryset below.
    deduction = (request.GET.get('deduction') or '').strip().upper()
    _valid_deductions = {v for v, _l in get_enabled_deduction_categories()}
    if deduction not in _valid_deductions:
        deduction = ''
    if deduction:
        loans = loans.filter(owner__deduction_category=deduction)

    # Employer drill-down: clicking an employer on a grouped report filters every
    # queryset to that employer so the report shows the individual client loans.
    drill_employer = (request.GET.get('employer') or '').strip()
    if drill_employer:
        loans = _blank_aware_filter(loans, 'owner__employer', drill_employer)

    # Period drill-down: clicking a period on a periodic disbursement report lists
    # the individual loans that make up that period's total.
    drill_period = (request.GET.get('period') or '').strip()

    # Group drill-down: clicking a group (region / product / status / client /
    # company) on a grouped report lists the individual loans in that group.
    drill_group = (request.GET.get('group') or '').strip()

    period_loans = loans.filter(funding_date__gte=start_date, funding_date__lte=end_date)
    statements = _statement_base().filter(date__gte=start_date, date__lte=end_date)
    payments = _payment_base().filter(date__gte=start_date, date__lte=end_date)
    if deduction:
        statements = statements.filter(loanref__owner__deduction_category=deduction)
        payments = payments.filter(owner__deduction_category=deduction)
    if drill_employer:
        statements = _blank_aware_filter(statements, 'loanref__owner__employer', drill_employer)
        payments = _blank_aware_filter(payments, 'owner__employer', drill_employer)
    description = 'Use the date filters to narrow the report period. Tables are limited to the first 500 rows where detailed records are shown.'
    # Slugs whose Employer column drills down to individual client loans.
    employer_grouped_slugs = {'loan-distribution-employer', 'disbursement-by-employer',
                              'loan-income-by-employer', 'expected-by-employer',
                              'alesco-per-employer', 'expected-vs-actual-repayments',
                              'payroll-deduction-alesco'}

    allocate_rows = None
    referrers_for_allocate = None
    if slug == 'loans-by-referrer' and drill_group == '— No referrer —':
        from referral.models import Referrer as _Ref
        _src = (loans.exclude(funded_category='ARCHIVED')
                .filter(owner__referred_by__isnull=True)
                .select_related('owner').order_by('owner__last_name'))
        columns = ['Ref', 'Client', 'Employer', 'Disbursed', 'Outstanding', 'Status']
        rows = []
        allocate_rows = []
        for l in _src:
            client = f'{l.owner.first_name} {l.owner.last_name}'.strip() if l.owner else 'Unknown'
            rows.append([_text(l.ref), client, _text(l.owner.employer if l.owner else None),
                         _money(l.amount), _money(l.total_outstanding), _text(l.status)])
            allocate_rows.append({'loan': l, 'client': client})
        referrers_for_allocate = _Ref.objects.filter(status='ACTIVE').order_by('name')
        description = ('Loans whose clients have NO referrer allocated. Pick a referrer against a client and '
                       'press Assign — the allocation saves immediately (and records the referral commission).')
        drill_group_handled = True
    else:
        drill_group_handled = False

    if not drill_group_handled and drill_group and slug in _GROUP_DRILL_FIELDS:
        gfield, only_outstanding = _GROUP_DRILL_FIELDS[slug]
        _src = period_loans if slug == 'disbursement-by-client' else loans
        _src = _blank_aware_filter(_src, gfield, drill_group)
        if only_outstanding:
            _src = _src.filter(total_outstanding__gt=0)
        columns, rows = _loan_rows(_src.select_related('owner').order_by('owner__last_name', 'funding_date'))
        description = 'Individual loans for: {}'.format(drill_group)
    elif slug in ['total-loan-disbursement', 'loan-disbursement-summary']:
        columns, rows = _periodic_loan_rows(period_loans, 'month')
    elif slug in _DISBURSEMENT_PERIODIC:
        trunc = _DISBURSEMENT_PERIODIC[slug]
        if drill_period:
            lo, hi = _period_bounds(drill_period, trunc)
            columns, rows = _loan_rows(
                period_loans.filter(funding_date__gte=lo, funding_date__lte=hi)
                .select_related('owner').order_by('funding_date', 'owner__last_name'))
            description = 'Individual loans disbursed {}.'.format(
                lo.strftime('%d %b %Y') if lo == hi else f"between {lo:%d %b %Y} and {hi:%d %b %Y}")
        else:
            columns, rows = _periodic_loan_rows(period_loans, trunc)
            description = 'Loans disbursed per {} in the selected window. Click a period to list the individual loans that make up its total.'.format(trunc)
    elif slug == 'outstanding-principal-interest':
        columns, rows = _loan_rows(loans.filter(total_outstanding__gt=0))
    elif slug == 'interest-earned-unearned':
        columns, rows = _interest_earned_unearned_rows(loans)
        description = 'Interest earned (already collected) versus unearned interest (still outstanding/receivable), by product — including default interest.'
    elif slug == 'active-vs-closed-loans':
        columns, rows = _grouped_loan_rows(loans, 'funded_category', 'Loan Status')
    elif slug in ['loan-distribution-employer', 'disbursement-by-employer', 'loan-income-by-employer']:
        _src = period_loans if slug == 'disbursement-by-employer' else loans
        if drill_employer:
            columns, rows = _loan_rows(_src)
            description = 'Individual client loans for {}.'.format(drill_employer)
        else:
            columns, rows = _grouped_loan_rows(_src, 'owner__employer', 'Employer')
    elif slug == 'expected-by-employer':
        if drill_employer:
            columns, rows = _expected_repayment_rows(loans.filter(next_payment_date__gte=start_date, next_payment_date__lte=end_date))
            description = 'Expected repayments for {} due in the selected pay window.'.format(drill_employer)
        else:
            columns, rows = _expected_by_employer_rows(loans.filter(next_payment_date__gte=start_date, next_payment_date__lte=end_date))
            description = 'How much to expect from each employer for repayments due in the selected pay window. Click an employer to see the individual client loans.'
    elif slug == 'loan-distribution-region':
        columns, rows = _grouped_loan_rows(loans, 'owner__residential_province', 'Region')
    elif slug == 'loan-distribution-product':
        columns, rows = _grouped_loan_rows(loans, 'loan_type', 'Product')
    elif slug == 'loan-income-by-product':
        columns, rows = _income_by_group_rows(loans, 'loan_type', 'Product')
        description = 'Income by product — interest and default interest collected plus processing fees, with the combined income total.'
    elif slug == 'loans-outstanding-company':
        columns, rows = _grouped_loan_rows(loans.filter(total_outstanding__gt=0), 'owner__employer', 'Department/Company')
    elif slug in ['disbursement-by-client', 'loan-income-by-client', 'client-outstanding-balance']:
        columns, rows = _grouped_loan_rows(loans, 'owner__email', 'Client')
    elif slug == 'new-vs-repeat-loans':
        columns, rows = _grouped_loan_rows(period_loans, 'classification', 'Classification')
    elif slug == 'processing-fees-summary':
        columns, rows = _periodic_loan_rows(period_loans, 'month')
    elif slug == 'processing-fees':
        # Detailed listing: one row per funded loan that attracted a processing
        # fee — these are the specific loans making up the total.
        columns, rows, _pf_total = _processing_fee_detail_rows(period_loans, 'month')
        exported = _maybe_export(request, report, slug, start_date, end_date, columns, rows)
        if exported:
            return exported
        context = {
            'nav': 'report_detail',
            'categories': REPORT_CATEGORIES,
            'report': report,
            'slug': slug,
            'description': 'Every funded loan in the selected period that attracted a processing fee — one row per loan, '
                           'making up the processing-fee total. Sortable; export to CSV/Excel/PDF.',
            'start_date': start_date,
            'end_date': end_date,
            'summary': [('Total Fees Collected', _money(_pf_total)), ('Loans with Fees', _number(len(rows)))],
            'columns': columns,
            'rows': rows,
            'totals': _compute_totals(columns, rows),
            'deduction': deduction,
            'deduction_categories': list(get_enabled_deduction_categories()),
            'row_count': len(rows),
            'ref_col': _ref_col(columns),
        }
        return render(request, 'report_detail.html', context)
    elif slug == 'expected-this-pay':
        columns, rows = _expected_repayment_rows(loans.filter(next_payment_date__gte=start_date, next_payment_date__lte=end_date))
        description = 'All client repayments expected (due) within the selected date window, sortable by any column, with the total expected at the bottom.'
    elif slug == 'expected-next-pay':
        next_start = end_date + datetime.timedelta(days=1)
        next_end = next_start + datetime.timedelta(days=14)
        columns, rows = _expected_repayment_rows(loans.filter(next_payment_date__gte=next_start, next_payment_date__lte=next_end))
        description = 'Client repayments expected in the next pay window (14 days after the selected end date), with the total at the bottom.'
    elif slug == 'payroll-deduction-alesco':
        columns, rows = _expected_vs_actual_rows(start_date, end_date, payroll_only=True, employer=drill_employer or None)
        if drill_employer:
            description = 'Payroll-deduction expected vs actual for {} (per client).'.format(drill_employer)
    elif slug == 'alesco-per-pay-period':
        columns, rows = _alesco_line_rows(start_date, end_date)
        description = ('Confirmed Alesco payroll deductions for the selected pay period(s), one row per client — '
                       'the amount paid this period against the expected scheduled repayment, with the variance. '
                       'Filter by the Period End date (single pay period) or a wider range. Sortable; export to PDF/Excel.')
    elif slug == 'alesco-per-employer':
        if drill_employer:
            columns, rows = _alesco_line_rows(start_date, end_date, employer=drill_employer)
            description = 'Confirmed Alesco payroll deductions for {} — one row per client (paid vs expected, with variance).'.format(drill_employer)
        else:
            columns, rows = _alesco_line_rows(start_date, end_date, group_by_employer=True)
            description = ('Confirmed Alesco payroll deductions grouped by employer for the selected pay period(s) — '
                           'total paid vs expected and the variance. Click an employer to drill into the individual clients. '
                           'Filter by pay period or a wider date range. Sortable; export to PDF/Excel.')
    elif slug == 'expected-vs-actual-repayments':
        columns, rows = _expected_vs_actual_rows(start_date, end_date, employer=drill_employer or None)
        if drill_employer:
            description = 'Expected (from schedule) vs actual received for {} — per client.'.format(drill_employer)
        else:
            description = ('Expected repayments (from the immutable schedule, so they do not clear once paid) vs '
                           'actual received, per employer. Click an employer to see individual clients — or use '
                           '"Expected vs Actual by Client" for a flat list of every loan account.')
    elif slug == 'loans-by-referrer' and allocate_rows is None:
        from admin1.models import referral_program_enabled
        if not referral_program_enabled():
            from django.contrib import messages as _msg
            from django.shortcuts import redirect as _redirect
            _msg.error(request, 'The referral programme is switched OFF (Referrers → Settings).', extra_tags='warning')
            return _redirect('reports')
        columns, rows = _loans_by_referrer_rows(loans.exclude(funded_category='ARCHIVED'))
        description = ('Loan portfolio grouped by referrer — how many clients and loans each referrer brought in, '
                       'the money disbursed and outstanding, and their quality: loans in default/arrears and the '
                       'percentage in good standing. Best-performing referrers sort to the top. '
                       'Click a referrer to see their individual loans.')
    elif slug == 'expected-vs-actual-by-client':
        columns, rows = _expected_vs_actual_rows(start_date, end_date, by_client=True)
        description = ('Every loan account\'s expected repayments (from the immutable schedule) vs actual received '
                       'in the window — one row per loan across all employers, so you can see at a glance who paid '
                       'and who did not (Actual K0.00). The Advance / (Arrears) column shows each client\'s CURRENT '
                       'standing — a positive figure is prepaid advance credit, a bracketed/negative figure is money '
                       'owed in arrears — so a client who has already defaulted this period is flagged even before '
                       'you report to the directors. Sortable by any column; totals at the bottom; '
                       'export to CSV/Excel/PDF.')
    elif slug in ['collections-weekly', 'collections-fortnightly', 'collections-monthly']:
        _trunc = {'collections-weekly': 'week', 'collections-fortnightly': 'fortnight',
                  'collections-monthly': 'month'}[slug]
        if drill_period:
            lo, hi = _period_bounds(drill_period, _trunc)
            _pmts = _payment_base().filter(date__gte=lo, date__lte=hi)
            if drill_employer:
                _pmts = _blank_aware_filter(_pmts, 'owner__employer', drill_employer)
            columns, rows, _ct = _collections_detail_rows(_pmts)
            description = 'Individual repayments collected {}.'.format(
                lo.strftime('%d %b %Y') if lo == hi else f'between {lo:%d %b %Y} and {hi:%d %b %Y}')
        else:
            columns, rows = _periodic_collection_rows(statements, _trunc, start_date)
            _label = {'week': 'week', 'fortnight': 'fortnight', 'month': 'month'}[_trunc]
            description = ('Repayments collected in the selected period, aggregated per {} — number of payments and '
                           'the principal, interest and default interest collected. Click a period to see the '
                           'individual payments that make it up.').format(_label)
    elif slug in ['collections-total', 'repayments-summary', 'interest-collected', 'recovery-collections']:
        if slug == 'recovery-collections':
            statements = statements.filter(loanref__funded_category='RECOVERY')
        columns, rows = _statement_collection_rows(statements)
    elif slug == 'arrears-0-30':
        columns, rows = _loan_rows(loans.filter(days_in_default__gte=0, days_in_default__lte=30, total_arrears__gt=0))
    elif slug == 'arrears-31-60':
        columns, rows = _loan_rows(loans.filter(days_in_default__gte=31, days_in_default__lte=60, total_arrears__gt=0))
    elif slug == 'arrears-61-90':
        columns, rows = _loan_rows(loans.filter(days_in_default__gte=61, days_in_default__lte=90, total_arrears__gt=0))
    elif slug == 'arrears-90-plus':
        columns, rows = _loan_rows(loans.filter(days_in_default__gt=90, total_arrears__gt=0))
    elif slug == 'non-performing-loans':
        columns, rows = _loan_rows(loans.filter(Q(days_in_default__gt=90) | Q(funded_category__in=['RECOVERY', 'BAD', 'WOFF'])))
    elif slug == 'defaulted-loans-report':
        columns, rows = _loan_rows(loans.filter(status='DEFAULTED'))
    elif slug == 'default-interest-report':
        columns, rows = _loan_rows(loans.filter(default_interest_receivable__gt=0))
    elif slug == 'missed-payroll-deductions':
        columns, rows = _loan_rows(loans.filter(next_payment_date__lt=timezone.localdate(), total_arrears__gt=0))
    elif slug == 'interest-accrued':
        columns, rows = _interest_accrued_rows(loans)
        description = 'Total interest accrued (charged over the loan life) by product, split into interest collected and interest still outstanding.'
    elif slug == 'client-loan-statement':
        columns, rows = _statement_collection_by_client_rows(statements)
        description = 'Per-client loan statement summary — payments collected on each loan in the period, split into principal, interest and default interest.'
    elif slug == 'loan-amortization-schedule':
        columns, rows = _loan_rows(loans.order_by('owner__email', 'expected_end_date'))
        description = 'Uses each loan repayment amount, repayment count, start date, next payment date, and expected end date as the amortization schedule source.'
    elif slug == 'client-repayment-history':
        columns, rows = _periodic_payment_rows(payments, 'month')
    elif slug == 'top-borrowers':
        columns, rows = _loan_rows(loans.order_by('-total_outstanding'))
    elif slug == 'repeat-borrowers':
        columns, rows = _loan_rows(loans.filter(Q(classification__in=['ADDITIONAL', 'REFINANCED']) | Q(owner__number_of_loans__gt=1)))
    elif slug == 'write-off-report':
        columns, rows = _loan_rows(loans.filter(funded_category__in=['WOFF', 'BAD']))
    elif slug == 'recovered-loans-report':
        columns, rows = _loan_rows(loans.filter(Q(recovery_date__isnull=False) | Q(funded_category='COMPLETED')))
    elif slug == 'outstanding-loan-report':
        columns, rows = _loan_rows(loans.filter(total_outstanding__gt=0))
    elif slug == 'provision-bad-debts':
        columns, rows = _grouped_loan_rows(loans.filter(Q(considered_unrecoverable__gt=0) | Q(funded_category__in=['BAD', 'WOFF'])), 'aging_category', 'Aging Category')
    elif slug == 'clients-by-deduction-category':
        columns, rows = _clients_by_deduction_category_rows()
        description = 'How many clients (and their active loan book) fall under each payroll/deduction category.'
    elif slug == 'business-performance':
        columns, rows = _business_performance_rows(start_date, end_date)
        description = ('A single-page management view of loan-book health (as of today) and business '
                       'performance for the selected period. Set the date filters to your reporting window.')
    elif slug in ['portfolio-report', 'collection-rate', 'default-rate', 'profitability-report', 'growth-trends']:
        columns, rows = _dashboard_rows(slug, start_date, end_date)
    elif slug in ['pfc-weekly', 'pfc-fortnightly', 'pfc-monthly', 'pfc-annually']:
        _trunc_map = {'pfc-weekly': 'week', 'pfc-fortnightly': 'fortnight', 'pfc-monthly': 'month', 'pfc-annually': 'year'}
        _trunc = _trunc_map[slug]
        columns, rows, _pfc_total = _periodic_fee_rows(period_loans, _trunc, start_date)
        exported = _maybe_export(request, report, slug, start_date, end_date, columns, rows)
        if exported:
            return exported
        _lbl = {'week': 'week', 'fortnight': 'fortnight', 'month': 'month', 'year': 'year'}[_trunc]
        context = {
            'nav': 'report_detail',
            'categories': REPORT_CATEGORIES,
            'report': report,
            'slug': slug,
            'description': ('Processing fees collected in the selected period, aggregated per {}. '
                            'For the loan-by-loan breakdown, use the “Processing Fees” report.').format(_lbl),
            'start_date': start_date,
            'end_date': end_date,
            'summary': [('Total Fees Collected', _money(_pfc_total)), ('Periods', _number(len(rows)))],
            'columns': columns,
            'rows': rows,
            'totals': _compute_totals(columns, rows),
            'deduction': deduction,
            'deduction_categories': list(get_enabled_deduction_categories()),
            'row_count': len(rows),
            'ref_col': _ref_col(columns),
        }
        return render(request, 'report_detail.html', context)
    else:
        columns, rows = _loan_rows(period_loans)

    exported = _maybe_export(request, report, slug, start_date, end_date, columns, rows)
    if exported:
        return exported

    context = {
        'nav': 'report_detail',
        'categories': REPORT_CATEGORIES,
        'report': report,
        'slug': slug,
        'description': description,
        'start_date': start_date,
        'end_date': end_date,
        'summary': _loan_summary(loans),
        'columns': columns,
        'rows': rows,
        'totals': _compute_totals(columns, rows),
        'row_count': len(rows),
        'employer_col': (columns.index('Employer') if (slug in employer_grouped_slugs and not drill_employer and 'Employer' in columns) else None),
        'period_col': (columns.index('Period') if ((slug in _DISBURSEMENT_PERIODIC or slug in ['collections-weekly', 'collections-fortnightly', 'collections-monthly']) and not drill_period and 'Period' in columns) else None),
        'group_col': (0 if (slug in _GROUP_DRILL_FIELDS and not drill_group and rows) else None),
        'allocate_rows': allocate_rows,
        'referrers_for_allocate': referrers_for_allocate,
        'ref_col': _ref_col(columns),
        'drill_employer': drill_employer,
        'drill_period': drill_period,
        'drill_group': drill_group,
        'deduction': deduction,
        'deduction_categories': list(get_enabled_deduction_categories()),
    }
    return render(request, 'report_detail.html', context)
