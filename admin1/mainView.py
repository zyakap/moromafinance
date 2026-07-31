import datetime
from django.utils import timezone
import re
from decimal import Decimal
from django.conf import settings
from django.contrib import messages
from django.shortcuts import redirect, render
from django.http import HttpResponse
from accounts.models import User, UserProfile, Bank, BankBranch
from accounts.forms import PersonalInfoForm, UserUploadForm, EmployerInfoUpdateForm, JobInfoUpdateForm, AddressInfoForm, BankAccountInfoForm
from loan.forms import PaymentForm as LoanPaymentForm, LoanApplicationForm, RequiredUploadForm
from loan.models import Loan, Statement, Payment, PaymentUploads
from .forms import AdminSettingsForm, DateSettingsForm, BankForm, BankBranchForm, EmployerForm, LoanSettingsForm, CreditCheckSettingsForm, DccSettingsForm, ProcessingFeeTierForm, RolesSettingsForm, TermsAndContractsSettingsForm, CreditAssessmentSettingsForm
from .models import AdminSettings, FormFieldSetting, Employer, ProcessingFeeTier, get_loan_config, get_bank_config, get_document_theme, calc_default_interest
from django.template.loader import render_to_string
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.utils.html import strip_tags
sender = settings.DEFAULT_SENDER_EMAIL

#Class Based Views
from django.views.generic.base import View

from django.db.models import Sum

from accounts.functions import admin_check, staff_or_admin_check

#########################
####   PAGES
#########################

@admin_check
def messages_admin(request):
    
    return render(request, 'messages_admin.html', {'nav': 'messages_admin'})

@admin_check
def support_system_admin(request):
    
    return render(request, 'support_system_admin.html', {'nav': 'support_system_admin'})

def admin_instructions(request):
    return render(request, 'admin_instructions.html', {'nav': 'admin_instructions'})


@admin_check
def global_search(request):
    from django.db.models import Q
    query = request.GET.get('q', '').strip()
    clients = []
    loans = []
    if query:
        client_q = (
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(nid_number__icontains=query) |
            Q(employer__icontains=query)
        )
        # mobile fields are IntegerField — only try numeric queries
        if query.isdigit():
            client_q |= Q(mobile1=int(query)) | Q(mobile2=int(query))
        clients = UserProfile.objects.filter(client_q).select_related('user').order_by('last_name', 'first_name')[:30]

        loans = Loan.objects.filter(
            Q(ref__icontains=query) |
            Q(status__icontains=query) |
            Q(loan_type__icontains=query) |
            Q(owner__first_name__icontains=query) |
            Q(owner__last_name__icontains=query) |
            Q(owner__user__email__icontains=query)
        ).select_related('owner', 'owner__user').order_by('-id')[:30]

    return render(request, 'admin_global_search.html', {
        'query': query,
        'clients': clients,
        'loans': loans,
        'nav': 'global_search',
    })

#########################
####   SETTINGS
#########################

@admin_check
def admin_dashboard(request):
    """Business-oversight dashboard — portfolio health, cash performance and
    action items, all from the shared metrics module (single source of truth)."""
    from admin1 import dashboard_metrics
    context = dashboard_metrics.build('admin')
    context['nav'] = 'admin_dashboard'

    # DCC running cost total (only when pay-per-view checks are enabled).
    # The bureau's billing summary is cached briefly so the dashboard never
    # waits on a slow/unreachable bureau on every load.
    from dcc.functions import dcc_enabled as _dcc_on, get_billing_summary
    context['dcc_enabled'] = _dcc_on()
    if context['dcc_enabled']:
        from django.core.cache import cache
        from dcc.models import DccViewLog
        summary = cache.get('dcc_billing_summary')
        if summary is None:
            summary = get_billing_summary() or {}
            cache.set('dcc_billing_summary', summary, 600)
        now = timezone.now()
        context['dcc_billing'] = summary or None
        context['dcc_views_month'] = DccViewLog.objects.filter(
            unlocked_at__year=now.year, unlocked_at__month=now.month).count()

    return render(request, 'admin_dashboard.html', context)

@admin_check
def admin_settings_general(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    if request.method == 'POST':
        form = AdminSettingsForm(request.POST, instance=setting1)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            from accounts.templatetags.finance_formats import invalidate_display_settings_cache
            invalidate_display_settings_cache()
            messages.success(request, "General settings have been updated.")
            return redirect('admin_settings_general')
    else:
        form = AdminSettingsForm(instance=setting1)

    return render(request, 'settings.html', {'form': form, 'nav': 'admin_settings_general'})


@admin_check
def admin_settings_appearance(request):
    from .forms import AppearanceSettingsForm
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = AdminSettings.objects.create(settings_name="setting1")

    if request.method == 'POST':
        form = AppearanceSettingsForm(request.POST, request.FILES, instance=setting1)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            messages.success(request, "Appearance settings have been updated. Refresh to see the changes.")
            return redirect('admin_settings_appearance')
    else:
        form = AppearanceSettingsForm(instance=setting1)

    from .models import get_appearance
    return render(request, 'settings_appearance.html', {
        'form': form, 'nav': 'admin_settings_appearance', 'appearance_now': get_appearance(),
    })


# Fields that originated as tenant-specific additions (not part of stock loanmasta).
# New tenants get these OFF by default — a tenant opts in via Settings > Form
# Settings if they want to collect them.
OPTIONAL_BY_DEFAULT_FIELDS = {
    ('personal_info', 'number_of_dependents'),
    ('personal_info', 'ages_of_dependents'),
    ('address_info', 'years_at_current_residence'),
    ('employer_info', 'immediate_supervisor'),
    ('bank_info', 'bank_bsb'),
    ('bank_info', 'bank_account_type'),
    ('loan_application', 'purpose_of_loan'),
    ('required_uploads', 'signed_contract'),
}


@admin_check
def admin_settings_form(request):
    form_config = [
        ('personal_info', 'Personal Information', PersonalInfoForm),
        ('address_info', 'Contact & Address', AddressInfoForm),
        ('bank_info', 'Bank Account', BankAccountInfoForm),
        ('user_uploads', 'User Uploads', UserUploadForm),
        ('employer_info', 'Employer Information', EmployerInfoUpdateForm),
        ('job_info', 'Job Information', JobInfoUpdateForm),
        ('payment', 'Payment Form', LoanPaymentForm),
        ('loan_application', 'Loan Application', LoanApplicationForm),
        ('required_uploads', 'Required Uploads', RequiredUploadForm),
    ]

    for form_name, form_label, form_class in form_config:
        for field_name, field in form_class.base_fields.items():
            default_enabled = (form_name, field_name) not in OPTIONAL_BY_DEFAULT_FIELDS
            FormFieldSetting.objects.get_or_create(
                form_name=form_name,
                field_name=field_name,
                defaults={'enabled': default_enabled, 'required': field.required if default_enabled else False}
            )

    if request.method == 'POST':
        for setting in FormFieldSetting.objects.all():
            enabled_key = f'enabled__{setting.form_name}__{setting.field_name}'
            required_key = f'required__{setting.form_name}__{setting.field_name}'
            setting.enabled = enabled_key in request.POST
            setting.required = required_key in request.POST
            setting.save()
        
        try:
            setting = AdminSettings.objects.get(settings_name='setting1')
        except:
            setting = AdminSettings.objects.create(settings_name='setting1')
        setting.employer_preregistration_required = 'employer_preregistration_required' in request.POST
        default_payment_description = request.POST.get('default_payment_description', '').strip()
        setting.default_payment_description = default_payment_description or 'Fortnightly Salary Deduction'
        setting.save()

        messages.success(request, 'Form settings have been updated.')
        return redirect('admin_settings_form')

    employer_preregistration_required = False
    default_payment_description = 'Fortnightly Salary Deduction'
    try:
        setting = AdminSettings.objects.get(settings_name='setting1')
        employer_preregistration_required = setting.employer_preregistration_required
        default_payment_description = setting.default_payment_description or 'Fortnightly Salary Deduction'
    except:
        pass

    blocks = []
    for form_name, form_label, form_class in form_config:
        settings_by_field = {
            setting.field_name: setting
            for setting in FormFieldSetting.objects.filter(form_name=form_name)
        }
        fields = []
        for field_name, field in form_class.base_fields.items():
            setting = settings_by_field.get(field_name)
            default_enabled = (form_name, field_name) not in OPTIONAL_BY_DEFAULT_FIELDS
            fields.append({
                'name': field_name,
                'label': field.label or field_name.replace('_', ' ').title(),
                'enabled': setting.enabled if setting else default_enabled,
                'required': setting.required if setting else (field.required if default_enabled else False),
            })
        block = {'name': form_name, 'label': form_label, 'fields': fields}
        if form_name == 'employer_info':
            block['preregistration'] = employer_preregistration_required
        if form_name == 'payment':
            block['default_payment_description'] = default_payment_description
        blocks.append(block)

    return render(request, 'settings_form.html', {'nav': 'admin_settings_form', 'blocks': blocks})

@admin_check
def admin_settings_date(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
        initial_data = {'date_format': setting1.date_format}
    except:
        initial_data = {'date_format': 'Y-m-d'}

    if request.method == 'POST':
        date_form = DateSettingsForm(request.POST)
        if date_form.is_valid():
            try:
                setting1 = AdminSettings.objects.get(settings_name="setting1")
            except:
                setting1 = AdminSettings.objects.create(settings_name="setting1")
            setting1.date_format = date_form.cleaned_data['date_format']
            setting1.save()
            from accounts.templatetags.finance_formats import invalidate_display_settings_cache
            invalidate_display_settings_cache()
            messages.success(request, "Date settings have been updated.")
            return redirect('admin_settings_date')
    else:
        date_form = DateSettingsForm(initial=initial_data)

    return render(request, 'settings_date.html', {'date_form': date_form, 'nav': 'admin_settings_date'})

@admin_check
def admin_settings_loans(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    tier_form = ProcessingFeeTierForm()
    edit_tier = None

    edit_tier_id = request.GET.get('edit_tier')
    if edit_tier_id:
        try:
            edit_tier = ProcessingFeeTier.objects.get(pk=edit_tier_id)
            tier_form = ProcessingFeeTierForm(instance=edit_tier)
        except ProcessingFeeTier.DoesNotExist:
            pass

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_loan_settings':
            try:
                setting1 = AdminSettings.objects.get(settings_name="setting1")
            except AdminSettings.DoesNotExist:
                setting1 = AdminSettings.objects.create(settings_name="setting1")
            loan_form = LoanSettingsForm(request.POST, instance=setting1)
            if loan_form.is_valid():
                loan_form.save()
                messages.success(request, "Loan settings have been updated.")
            return redirect('admin_settings_loans')

        elif action == 'add_tier':
            tier_form = ProcessingFeeTierForm(request.POST)
            if tier_form.is_valid():
                tier_form.save()
                messages.success(request, 'Processing fee tier added.')
                return redirect('admin_settings_loans')

        elif action == 'edit_tier':
            tier_id = request.POST.get('tier_id')
            try:
                edit_tier = ProcessingFeeTier.objects.get(pk=tier_id)
                tier_form = ProcessingFeeTierForm(request.POST, instance=edit_tier)
                if tier_form.is_valid():
                    tier_form.save()
                    messages.success(request, 'Processing fee tier updated.')
                    return redirect('admin_settings_loans')
            except ProcessingFeeTier.DoesNotExist:
                messages.error(request, 'Tier not found.')

        elif action == 'delete_tier':
            tier_id = request.POST.get('tier_id')
            try:
                ProcessingFeeTier.objects.get(pk=tier_id).delete()
                messages.success(request, 'Processing fee tier deleted.')
            except ProcessingFeeTier.DoesNotExist:
                messages.error(request, 'Tier not found.')
            return redirect('admin_settings_loans')

        elif action == 'toggle_tier':
            tier_id = request.POST.get('tier_id')
            try:
                tier = ProcessingFeeTier.objects.get(pk=tier_id)
                tier.active = not tier.active
                tier.save()
                messages.success(request, f"Tier {'activated' if tier.active else 'deactivated'}.")
            except ProcessingFeeTier.DoesNotExist:
                messages.error(request, 'Tier not found.')
            return redirect('admin_settings_loans')

    else:
        loan_form = LoanSettingsForm(instance=setting1)

    tiers = ProcessingFeeTier.objects.all().order_by('min_amount')
    return render(request, 'settings_loan.html', {
        'loan_form': loan_form,
        'tier_form': tier_form,
        'edit_tier': edit_tier,
        'tiers': tiers,
        'nav': 'admin_settings_loans',
    })

@admin_check
def admin_settings_credit(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    if request.method == 'POST':
        credit_form = CreditCheckSettingsForm(request.POST, instance=setting1)
        if credit_form.is_valid():
            obj = credit_form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            messages.success(request, "Credit Check settings have been updated.")
            return redirect('admin_settings_credit')
    else:
        credit_form = CreditCheckSettingsForm(instance=setting1)

    return render(request, 'settings_credit.html', {
        'credit_form': credit_form,
        'nav': 'admin_settings_credit',
    })


@admin_check
def admin_settings_dcc(request):
    """DCC Settings: enable/disable billed pay-per-view credit checks and the
    benchmark-score automation. The data feed to DCC is always on and is shown
    here for transparency only."""
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    if request.method == 'POST':
        dcc_form = DccSettingsForm(request.POST, instance=setting1)
        if dcc_form.is_valid():
            obj = dcc_form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            messages.success(request, "DCC settings have been updated.")
            return redirect('admin_settings_dcc')
    else:
        dcc_form = DccSettingsForm(instance=setting1)

    dcc_enabled = not (setting1 is not None and setting1.dcc_enabled == 'NO')

    return render(request, 'settings_dcc.html', {
        'dcc_form': dcc_form,
        'nav': 'admin_settings_dcc',
        'dcc_enabled': dcc_enabled,
        'dcc_endpoint': settings.DCC_ENDPOINT,
        'dcc_luid': settings.LUID,
        'dcc_key_set': bool(getattr(settings, 'DCC_API_KEY', '')),
    })

@admin_check
def admin_settings_roles(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    if request.method == 'POST':
        roles_form = RolesSettingsForm(request.POST, instance=setting1)
        if roles_form.is_valid():
            obj = roles_form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            messages.success(request, "Roles & Access settings have been updated.")
            return redirect('admin_settings_roles')
    else:
        roles_form = RolesSettingsForm(instance=setting1)

    return render(request, 'settings_roles.html', {
        'roles_form': roles_form,
        'nav': 'admin_settings_roles',
    })

@admin_check
def admin_settings_tc_contracts(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    if request.method == 'POST':
        tc_form = TermsAndContractsSettingsForm(request.POST, instance=setting1)
        if tc_form.is_valid():
            obj = tc_form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            messages.success(request, "T&C and Contracts settings have been updated.")
            return redirect('admin_settings_tc_contracts')
    else:
        tc_form = TermsAndContractsSettingsForm(instance=setting1)

    return render(request, 'settings_tc_contracts.html', {
        'tc_form': tc_form,
        'nav': 'admin_settings_tc_contracts',
    })

@admin_check
def admin_settings_credit_assessment(request):
    try:
        setting1 = AdminSettings.objects.get(settings_name="setting1")
    except AdminSettings.DoesNotExist:
        setting1 = None

    if request.method == 'POST':
        ca_form = CreditAssessmentSettingsForm(request.POST, instance=setting1)
        if ca_form.is_valid():
            obj = ca_form.save(commit=False)
            obj.settings_name = 'setting1'
            obj.save()
            messages.success(request, "Credit Assessment settings have been updated.")
            return redirect('admin_settings_credit_assessment')
    else:
        ca_form = CreditAssessmentSettingsForm(instance=setting1)

    return render(request, 'settings_credit_assessment.html', {
        'ca_form': ca_form,
        'nav': 'admin_settings_credit_assessment',
    })

@admin_check
def locations(request):
    
    return render(request, 'locations.html')

def _get_las_settings():
    """Return (date_fmt, col_flags_dict) from AdminSettings."""
    try:
        s = AdminSettings.objects.get(settings_name='setting1')
        date_fmt = s.date_format or 'd M, Y'
        cols = {
            'period':    s.las_col_period,
            'date':      s.las_col_date,
            'opening':   s.las_col_opening,
            'interest':  s.las_col_interest,
            'principal': s.las_col_principal,
            'repayment': s.las_col_repayment,
            'closing':   s.las_col_closing,
        }
        return date_fmt, cols
    except Exception:
        return 'd M, Y', {k: True for k in ('period','date','opening','interest','principal','repayment','closing')}


def _fmt_date(date_obj, django_fmt):
    """Convert a Django date format string to a Python strftime format and format the date."""
    import datetime
    if date_obj is None:
        return ''
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.date.fromisoformat(date_obj)
        except ValueError:
            return date_obj
    _map = {
        'Y': '%Y', 'y': '%y',
        'm': '%m', 'n': '%-m',
        'd': '%d', 'j': '%-d',
        'M': '%b', 'F': '%B',
        'D': '%a', 'l': '%A',
    }
    py_fmt = ''
    for ch in django_fmt:
        py_fmt += _map.get(ch, ch)
    return date_obj.strftime(py_fmt)


def _build_las_context(loan, user, domain):
    """Build the amortization schedule rows and totals for LAS PDF/email."""
    from decimal import Decimal
    import datetime

    date_fmt, cols = _get_las_settings()

    n = loan.number_of_fortnights or 0
    repayment = Decimal(str(loan.repayment_amount or 0))
    total_loan = Decimal(str(loan.total_loan_amount or 0))
    interest_total = Decimal(str(loan.interest or 0))

    # Get stored repayment dates (list of 'YYYY-MM-DD' strings)
    repayment_dates = loan.get_repayment_dates() or []
    # Optional per-fortnight repayment amounts (non-uniform schedules, e.g. a
    # CONCURRENT refinance where the repayment steps down). When present, the
    # Repayment column and closing-balance walk follow this list.
    repayment_amounts = loan.get_repayment_amounts() or []
    has_amounts = len(repayment_amounts) == n

    schedule = []
    opening = total_loan
    sum_interest = Decimal('0')
    sum_principal = Decimal('0')
    sum_repayment = Decimal('0')

    for i in range(1, n + 1):
        is_last = (i == n)
        # The balance is gross (principal + interest), so each repayment reduces it
        # by its full amount. The interest/principal split is informational and is
        # attributed in proportion to each period's repayment — this equals the flat
        # interest_total/n for a uniform schedule and stays correct (no negative
        # principal) for a stepped concurrent schedule. The final row clears whatever
        # remains, absorbing rounding.
        if is_last:
            repayment_row = opening
            closing = Decimal('0')
        else:
            repayment_row = Decimal(str(repayment_amounts[i - 1])) if has_amounts else repayment
            closing = round(opening - repayment_row, 2)

        interest_row = round(interest_total * (repayment_row / total_loan), 2) if total_loan else Decimal('0')
        principal = repayment_row - interest_row

        sum_interest += interest_row
        sum_principal += principal
        sum_repayment += repayment_row

        date_str = repayment_dates[i - 1] if i - 1 < len(repayment_dates) else ''
        schedule.append({
            'period': i,
            'date': _fmt_date(date_str, date_fmt),
            'opening': opening,
            'interest': round(interest_row, 2),
            'principal': round(principal, 2),
            'repayment': round(repayment_row, 2),
            'closing': closing,
        })
        opening = closing

    return {
        'loan': loan,
        'user': user,
        'domain': domain,
        'today': _fmt_date(datetime.date.today(), date_fmt),
        'schedule': schedule,
        'total_interest': round(sum_interest, 2),
        'total_principal': round(sum_principal, 2),
        'total_repayment': round(sum_repayment, 2),
        'cols': cols,
        'bank_config': get_bank_config(),
        'document_theme': get_document_theme(),
    }


# ── Document view / download helpers ───────────────────────────────────────
# CDN, LAS and Statement are shown on-screen (inline PDF preview) first, then the
# user chooses Download PDF or Download Excel. Each document view branches on the
# ?format= query parameter: absent → on-screen preview page; 'pdf' → PDF (inline
# when &inline=1, else attachment); 'xlsx' → Excel export.

def _document_preview_response(request, doc_title, has_excel=True, send_url=None):
    return render(request, 'document_preview.html', {
        'nav': '', 'doc_title': doc_title, 'doc_url': request.path, 'has_excel': has_excel,
        'send_url': send_url,
    })


def _xlsx_response(filename, title, headers, rows, meta=None):
    """Build a simple .xlsx download from headers + rows (+ optional meta pairs)."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Document'
    ws.append([title]); ws['A1'].font = Font(bold=True, size=14)
    for k, v in (meta or []):
        ws.append([k, v])
    ws.append([])
    if headers:
        ws.append(headers)
        theme_color = get_document_theme()['topbar_color'].lstrip('#')
        hdr_fill = PatternFill(fill_type='solid', start_color=theme_color, end_color=theme_color)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = hdr_fill
    for r in rows:
        ws.append([float(x) if isinstance(x, Decimal) else x for x in r])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


@staff_or_admin_check
def download_las(request, loan_ref):
    """View/Download the Loan Amortization Schedule (preview → PDF/Excel)."""
    fmt = request.GET.get('format')
    if not fmt:
        from django.urls import reverse
        return _document_preview_response(request, f'Loan Amortization Schedule — {loan_ref}', has_excel=True,
                                          send_url=reverse('send_doc_client', args=[loan_ref, 'las']))

    loan = Loan.objects.get(ref=loan_ref)
    user = UserProfile.objects.get(id=loan.owner_id)
    context = _build_las_context(loan, user, settings.DOMAIN)

    if fmt == 'xlsx':
        headers = ['Period', 'Date', 'Opening', 'Interest', 'Principal', 'Repayment', 'Closing']
        rows = [[r['period'], r['date'], r['opening'], r['interest'], r['principal'], r['repayment'], r['closing']]
                for r in context['schedule']]
        rows.append(['TOTAL', '', '', context['total_interest'], context['total_principal'], context['total_repayment'], ''])
        return _xlsx_response(f'LAS_{loan_ref}', f'Loan Amortization Schedule — {loan_ref}', headers, rows,
                              meta=[('Client', f"{user.first_name} {user.last_name}"), ('Loan Ref', loan_ref)])

    from moromafinance.pdf import django_pdf_response
    return django_pdf_response(
        request, 'custom/loan_amortization_schedule.html', context,
        f'LAS_{loan_ref}.pdf', inline=(request.GET.get('inline') == '1'),
    )


@staff_or_admin_check
def download_cdn(request, loan_ref):
    """View/Download the Cash Disbursement Notice (preview → PDF/Excel)."""
    import datetime

    from decimal import Decimal as _D

    fmt = request.GET.get('format')
    if not fmt:
        from django.urls import reverse
        return _document_preview_response(request, f'Cash Disbursement Notice — {loan_ref}', has_excel=True,
                                          send_url=reverse('send_doc_client', args=[loan_ref, 'cdn']))

    loan = Loan.objects.get(ref=loan_ref)
    user = UserProfile.objects.get(id=loan.owner_id)

    funding_date = loan.funding_date.strftime('%d/%m/%Y') if loan.funding_date else datetime.date.today().strftime('%d/%m/%Y')
    first_repayment_date = loan.next_payment_date.strftime('%d %B %Y') if loan.next_payment_date else ''

    # Amount actually disbursed to the client. If the processing fee is withheld
    # from the payout, the disbursed amount is the loan amount less the fee;
    # otherwise the full loan amount is paid out. (download_cdn previously omitted
    # this, so the CDN showed "Loan Amount Disbursed: K0.00".)
    try:
        _ls = AdminSettings.objects.get(settings_name='setting1')
        _fee_mode = getattr(_ls, 'processing_fee_collection_mode', 'CASH')
    except Exception:
        _fee_mode = 'CASH'
    _fee = _D(str(loan.processing_fee or 0))
    _loan_amount = _D(str(loan.amount or 0))
    _disbursed_amount = (_loan_amount - _fee) if (_fee_mode == 'WITHHELD' and _fee > 0) else _loan_amount

    bank_config = get_bank_config()
    context = {
        'loan': loan,
        'user': user,
        'today': datetime.date.today().strftime('%d %B %Y'),
        'funding_date': funding_date,
        'first_repayment_date': first_repayment_date,
        'disbursed_amount': _disbursed_amount,
        'fee_collection_mode': _fee_mode,
        'domain': settings.DOMAIN,
        'bank_name': bank_config["bank"],
        'bank_account_name': bank_config["bank_acc_name"].title(),
        'bank_account_number': bank_config["bank_acc_num"],
        'bank_branch': bank_config["bank_branch"].title(),
        'deduction_code': bank_config["alesco_dc"],
        'document_theme': get_document_theme(),
    }
    if fmt == 'xlsx':
        headers = ['Field', 'Value']
        rows = [
            ['Client', f"{user.first_name} {user.last_name}"],
            ['Loan Ref', loan_ref],
            ['Funding Date', funding_date],
            ['Loan Amount', _D(str(loan.amount or 0))],
            ['Processing Fee', _fee],
            ['Amount Disbursed', _disbursed_amount],
            ['Fee Collection Mode', _fee_mode],
            ['First Repayment Date', first_repayment_date],
            ['Alesco Deduction Code', context['deduction_code']],
            ['Bank', context['bank_name']],
            ['Account Name', context['bank_account_name']],
            ['Account Number', context['bank_account_number']],
            ['Branch', context['bank_branch']],
        ]
        return _xlsx_response(f'CDN_{loan_ref}', f'Cash Disbursement Notice — {loan_ref}', headers, rows)

    from moromafinance.pdf import django_pdf_response
    return django_pdf_response(
        request, 'custom/cash_disbursement_notice.html', context,
        f'CDN_{loan_ref}.pdf', inline=(request.GET.get('inline') == '1'),
    )


class DownloadApplicationByAdmin(View):

    template = 'custom/client_statement.html'
    
    def get(self, request, *args, **kwargs):
        
        loan_ref = self.kwargs['loanref']
        ##loan_ref = 'iBX1ZY264'
        loan = Loan.objects.get(ref=loan_ref)
        domain = settings.DOMAIN
        uid = loan.owner_id
        user = UserProfile.objects.get(pk=uid)
        usr = User.objects.get(pk=user.user_id)

        statements = Statement.objects.filter(loanref=loan)
        
        now = datetime.datetime.now()
        today = now.strftime("%d/%m/%Y")

        
        last_name_s = user.last_name[-1]

        from loan.functions import statement_summary
        data = {'loan':loan, 'user':user, 'usr': usr, 'last_name_s':last_name_s, 'domain': domain, 'statements': statements, 'today':today, 'bank_config': get_bank_config(), 'document_theme': get_document_theme() }
        data.update(statement_summary(loan, statements))

        from moromafinance.pdf import django_pdf_response
        return django_pdf_response(
            request, self.template, data, f'{loan.ref}.pdf'
        )



@admin_check
def make_default(request, loan_ref):
    
    if not request.user.is_authenticated:
        return redirect('login_user')
    
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.", extra_tags="danger")
        return redirect( 'dashboard')
    
    loan = Loan.objects.get(ref=loan_ref) 
    ref = loan
    date = loan.next_payment_date
    amount = loan.repayment_amount
    balance = loan.total_outstanding
    arrears = loan.total_arrears
    
    stat = Statement.objects.create(loanref=ref, date=date, default_amount=amount, statement='DEFAULT')
    stat.uid = loan.owner.uid
    stat.luid = settings.LUID
    stat.save()
        
    all_statements = Statement.objects.filter(loanref=loan).all().count()     
    stat.s_count = all_statements
    stat.ref = f'{loan_ref}SD{stat.s_count}' 
    stat.save()
    
    from loan.engine import default_interest_for
    int_on_def = default_interest_for(loan, amount, get_loan_config())
    new_balance = balance + int_on_def
    new_arrears = arrears + amount
    
    stat.interest_on_default = int_on_def
    stat.balance = new_balance
    stat.arrears = new_arrears
    stat.save()
    
    loan.last_default_date = date
    last_default_amount = amount
    loan.number_of_defaults += 1
    loan.total_arrears = new_arrears
    loan.default_interest_receivable += int_on_def
    loan.total_outstanding = new_balance
    loan.next_payment_date = date + datetime.timedelta(days=14)
    loan.category = 'CURRENT-DEFAULTED'
    loan.status = 'DEFAULTED'
    loan.save()
    
    messages.error(request, 'DEFAULT UPDATED SUCCESSFULLY', extra_tags='info')
    
    return redirect('loans')

from django.utils.decorators import method_decorator


@method_decorator(staff_or_admin_check, name='dispatch')
class DownloadLoanStatement(View):
    
    template = 'custom/client_statement.html'
    
    def get(self, request, *args, **kwargs):
        domain = settings.DOMAIN
        loan_ref = self.kwargs['loanref']

        fmt = request.GET.get('format')
        if not fmt:
            from django.urls import reverse
            return _document_preview_response(request, f'Loan Statement — {loan_ref}', has_excel=True,
                                              send_url=reverse('send_doc_client', args=[loan_ref, 'statement']))

        loan = Loan.objects.get(ref=loan_ref)

        uid = loan.owner_id
        user = UserProfile.objects.get(pk=uid)
        usr = User.objects.get(pk=user.user_id)

        last_name_s = user.last_name[-1]

        today = datetime.date.today().strftime('%x')

        statements = Statement.objects.filter(loanref=loan)

        if fmt == 'xlsx':
            headers = ['Date', 'Description', 'Debit', 'Credit', 'Principal', 'Interest', 'D.Interest', 'Arrears', 'Balance']
            rows = [[
                s.date, s.statement, s.debit, s.credit, s.principal_collected,
                s.interest_collected, s.default_interest_collected, s.arrears, s.balance,
            ] for s in statements]
            return _xlsx_response(f'{loan_ref}-Statement', f'Loan Statement — {loan_ref}', headers, rows,
                                  meta=[('Client', f"{user.first_name} {user.last_name}"), ('Loan Ref', loan_ref)])

        from loan.functions import statement_summary
        data = {'loan':loan, 'user':user, 'usr': usr, 'last_name_s':last_name_s, 'statements': statements, 'domain': domain, 'today': today, 'bank_config': get_bank_config(), 'document_theme': get_document_theme() }
        data.update(statement_summary(loan, statements))

        from moromafinance.pdf import django_pdf_response
        return django_pdf_response(
            request, self.template, data, f'{loan.ref}-Statement.pdf',
            inline=(request.GET.get('inline') == '1'),
        )
    
@admin_check
def reports(request):  

    return render(request, 'reports.html', {'nav': 'reports'})



@admin_check
def transactions(request):
    
    transactions = Statement.objects.prefetch_related('loanref','owner').order_by('-date')
   
    return render(request, 'transactions.html', { 'transactions': transactions })

@admin_check
def statements(request):

    transactions = Statement.objects.prefetch_related('loanref','owner').filter(type='PAYMENT').order_by('-date')
    
    return render(request, 'statements.html', { 'transactions': transactions })

@admin_check
def payments(request):

    transactions = Payment.objects.prefetch_related('loanref','owner').order_by('-date')
    
    return render(request, 'payments.html', { 'transactions': transactions })

@admin_check
def defaults(request):

    transactions = Statement.objects.prefetch_related('loanref','owner').filter(type='DEFAULT').order_by('-date')
    
    return render(request, 'defaults.html', { 'transactions': transactions })

@admin_check
def payment_uploads(request):

    payment_uploads = PaymentUploads.objects.prefetch_related('owner','loan').filter(status="UPLOADED")
    completed_uploads = PaymentUploads.objects.prefetch_related('owner','loan').filter(status="PROCESSED")

    return render(request, 'payment_uploads.html', {'nav': 'payment_uploads', 'payment_uploads': payment_uploads, 'completed_uploads': completed_uploads})

@admin_check
def process_upload(request, ref):

    payment_upload = PaymentUploads.objects.get(ref=ref)
    
    payment_upload.status = 'PROCESSED'
    payment_upload.save()
    messages.success(request, f"Payment upload {ref} has been processed.")

    payment_uploads = PaymentUploads.objects.prefetch_related('owner','loan').filter(status="UPLOADED")
    completed_uploads = PaymentUploads.objects.prefetch_related('owner','loan').filter(status="PROCESSED")


    return render(request, 'payment_uploads.html', {'nav': 'payment_uploads', 'payment_uploads': payment_uploads, 'completed_uploads': completed_uploads})

@admin_check
def admin_run_defaults(request):
    """Batch-process overdue defaults via the loan engine (advance netting,
    shortfall-based interest, immutable-schedule cursor). Honours the mercy
    period: a repayment is only defaulted once it is more than
    ``mercy_days`` past due."""
    from django.db import transaction
    from django.db.models import Q
    from loan.engine import create_default_for_loan

    mercy = get_loan_config().get('mercy_days') or 0
    loans = (Loan.objects.filter(Q(status='RUNNING') | Q(status='DEFAULTED'), category='FUNDED')
             .exclude(funded_category__in=['COMPLETED', 'ARCHIVED']))
    created = 0
    affected = 0
    for loan in loans:
        loan_defaulted = False
        while True:
            with transaction.atomic():
                locked = Loan.objects.select_for_update().get(pk=loan.pk)
                stat, ok, msg = create_default_for_loan(locked, mercy_days=mercy)
            if not ok:
                break
            created += 1
            loan_defaulted = True
        if loan_defaulted:
            affected += 1

    if created:
        messages.success(request, f'Created {created} default(s) across {affected} loan(s).', extra_tags='info')
    else:
        messages.info(request, 'No overdue repayments found to default (mercy period respected).', extra_tags='info')
    return redirect('admin_dashboard')
 

    
def _loan_view_redirect_name(request):
    """'view_loan' (admin portal) for superusers/ADMIN-category users, else
    'view_loan_staff' — so an action triggered from the staff portal sends
    staff back to a page they're actually allowed to see, instead of the
    admin-only view_loan (which would 403 them right after the action
    they just performed succeeded)."""
    if request.user.is_superuser:
        return 'view_loan'
    profile = UserProfile.objects.filter(user_id=request.user.id).first()
    if profile is not None and profile.category == 'ADMIN':
        return 'view_loan'
    return 'view_loan_staff'


@staff_or_admin_check
def create_default(request, loan_ref):
    from django.db import transaction
    from loan.engine import create_default_for_loan, preview_default

    redirect_name = _loan_view_redirect_name(request)

    # The engine only ever defaults the earliest scheduled repayment that is
    # strictly in the past, nets any advance balance, and charges interest on the
    # shortfall (or full amount) per the admin setting. It refuses to default a
    # repayment that is not yet due, so future repayments are never defaulted.
    #
    # If the client's advance balance would fully cover the missed repayment,
    # steer the operator to "Apply Advance to Missed Payments" instead of
    # silently defaulting — a default should only ever be charged on a genuine
    # shortfall.
    loan = Loan.objects.get(ref=loan_ref)
    due, shortfall, advance_applied = preview_default(loan)
    if due is not None and shortfall is not None and shortfall <= 0 and advance_applied and advance_applied > 0:
        messages.warning(
            request,
            f'{loan.ref} has an advance balance of K{loan.advance_balance:.2f} that fully covers the '
            f'repayment due {due.isoformat()}. Use "Apply Advance to Missed Payments" instead of Create Default.',
            extra_tags='warning',
        )
        return redirect(redirect_name, loan.ref)

    with transaction.atomic():
        loan = Loan.objects.select_for_update().get(ref=loan_ref)
        stat, ok, msg = create_default_for_loan(loan)

    if ok:
        messages.success(request, msg, extra_tags='info')
    else:
        messages.warning(request, msg, extra_tags='warning')

    return redirect(redirect_name, loan.ref)


@staff_or_admin_check
def reverse_last_statement(request, loan_ref):
    """Reverse the most recent statement line on a loan — undoing the WHOLE
    transaction that created it (payment + sibling statement lines + every
    loan-field effect), reconstructed from the ledger by loan.reversal.

    Reverses one line per click, newest first, and refuses once only the two
    structural funding lines (Loan Funded / Loan Interest Charged) remain — those
    and any REFINANCE line are never reversible. POST only, so it can't fire on
    an accidental navigation the way the old GET Create Default link did.
    """
    from loan.reversal import reverse_statement, ReversalError, REVERSIBLE_TYPES

    redirect_name = _loan_view_redirect_name(request)
    if request.method != 'POST':
        return redirect(redirect_name, loan_ref)

    loan = Loan.objects.get(ref=loan_ref)
    newest = Statement.objects.filter(loanref=loan).order_by('-pk').first()
    if newest is None:
        messages.warning(request, 'This loan has no statement lines to reverse.', extra_tags='warning')
        return redirect(redirect_name, loan_ref)
    if newest.type not in REVERSIBLE_TYPES:
        messages.warning(
            request,
            'Only the funding and interest-charged lines remain — these cannot be reversed.',
            extra_tags='warning',
        )
        return redirect(redirect_name, loan_ref)

    try:
        summary = reverse_statement(newest)
        messages.success(request, f'Last statement reversed. {summary}', extra_tags='info')
    except ReversalError as exc:
        messages.error(request, f'Could not reverse the last statement — {exc}', extra_tags='danger')

    return redirect(redirect_name, loan_ref)


@staff_or_admin_check
def apply_advance_to_missed_payment(request, loan_ref):
    """Settle the earliest overdue repayment from the client's advance
    balance instead of charging a default, when the advance fully covers it.

    Delegates to create_default_for_loan — the engine already nets any
    advance balance against the shortfall before deciding whether to default,
    so when the advance covers the repayment in full the engine settles the
    fortnight and records an ADVANCE-APPLIED statement with no default. This
    action exists so operators have an explicit, deliberate button for that
    outcome instead of relying on Create Default's silent netting.
    """
    from django.db import transaction
    from loan.engine import create_default_for_loan

    redirect_name = _loan_view_redirect_name(request)
    loan = Loan.objects.get(ref=loan_ref)
    if not loan.advance_balance or loan.advance_balance <= 0:
        messages.info(request, f'{loan.ref} has no advance balance to apply.', extra_tags='info')
        return redirect(redirect_name, loan.ref)

    with transaction.atomic():
        loan = Loan.objects.select_for_update().get(ref=loan_ref)
        stat, ok, msg = create_default_for_loan(loan)

    if ok:
        messages.success(request, msg, extra_tags='info')
    else:
        messages.warning(request, msg, extra_tags='warning')

    return redirect(redirect_name, loan.ref)


@staff_or_admin_check
def fix_loan_defaults(request, loan_ref):
    """Correction tool: reverse default lines that were created for dates that
    are NOT yet overdue (the old engine's future-dated-default bug), roll back
    their financial effect, and put those dates back into the schedule so they
    default correctly only once genuinely overdue."""
    from django.db import transaction
    from loan.engine import _D
    import datetime as _dt

    loan = Loan.objects.get(ref=loan_ref)
    today = _dt.date.today()
    future_defaults = Statement.objects.filter(loanref=loan, type='DEFAULT', date__gt=today).order_by('date')

    if request.method == 'POST':
        with transaction.atomic():
            loan = Loan.objects.select_for_update().get(ref=loan_ref)
            bad = list(Statement.objects.filter(loanref=loan, type='DEFAULT', date__gt=today).order_by('date'))
            readded = []
            for s in bad:
                di = _D(s.default_interest if s.default_interest is not None else s.credit)
                sf = _D(s.default_amount)
                loan.total_outstanding = _D(loan.total_outstanding) - di
                loan.default_interest_receivable = _D(loan.default_interest_receivable) - di
                loan.total_arrears = _D(loan.total_arrears) - sf
                loan.number_of_defaults = max(0, (loan.number_of_defaults or 0) - 1)
                readded.append(s.date.isoformat())
                s.delete()

            # Roll the immutable-schedule cursor back one slot per reversed
            # default so those fortnights become due (and next_payment_date is
            # re-derived). The schedule itself is not mutated.
            from loan import schedule as _sched
            _sched.unsettle(loan, len(bad))
            if _D(loan.total_arrears) <= 0:
                loan.total_arrears = 0
                if loan.status == 'DEFAULTED':
                    loan.status = 'RUNNING'
            loan.save()

        messages.success(request, f'Reversed {len(bad)} future-dated default(s) for {loan_ref} and restored their dates to the schedule.', extra_tags='info')
        return redirect('view_loan', loan_ref)

    return render(request, 'fix_loan_defaults.html', {
        'loan': loan, 'future_defaults': future_defaults, 'today': today,
    })


@admin_check
def admin_settings_banks(request):
    bank_form = BankForm()
    branch_form = BankBranchForm()
    edit_bank = None
    edit_branch = None

    edit_bank_id = request.GET.get('edit_bank')
    if edit_bank_id:
        try:
            edit_bank = Bank.objects.get(pk=edit_bank_id)
            bank_form = BankForm(instance=edit_bank)
        except Bank.DoesNotExist:
            pass

    edit_branch_id = request.GET.get('edit_branch')
    if edit_branch_id:
        try:
            edit_branch = BankBranch.objects.get(pk=edit_branch_id)
            branch_form = BankBranchForm(instance=edit_branch)
        except BankBranch.DoesNotExist:
            pass

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_bank':
            bank_form = BankForm(request.POST)
            if bank_form.is_valid():
                bank_form.save()
                messages.success(request, 'Bank added successfully.')
                return redirect('admin_settings_banks')

        elif action == 'edit_bank':
            bank_id = request.POST.get('bank_id')
            try:
                edit_bank = Bank.objects.get(pk=bank_id)
                bank_form = BankForm(request.POST, instance=edit_bank)
                if bank_form.is_valid():
                    bank_form.save()
                    messages.success(request, 'Bank updated successfully.')
                    return redirect('admin_settings_banks')
            except Bank.DoesNotExist:
                messages.error(request, 'Bank not found.')

        elif action == 'add_branch':
            branch_form = BankBranchForm(request.POST)
            if branch_form.is_valid():
                branch_form.save()
                messages.success(request, 'Branch added successfully.')
                return redirect('admin_settings_banks')

        elif action == 'edit_branch':
            branch_id = request.POST.get('branch_id')
            try:
                edit_branch = BankBranch.objects.get(pk=branch_id)
                branch_form = BankBranchForm(request.POST, instance=edit_branch)
                if branch_form.is_valid():
                    branch_form.save()
                    messages.success(request, 'Branch updated successfully.')
                    return redirect('admin_settings_banks')
            except BankBranch.DoesNotExist:
                messages.error(request, 'Branch not found.')

        elif action == 'toggle_bank':
            bank_id = request.POST.get('bank_id')
            try:
                bank = Bank.objects.get(pk=bank_id)
                bank.active = not bank.active
                bank.save()
                messages.success(request, f"Bank {'activated' if bank.active else 'deactivated'}.")
            except Bank.DoesNotExist:
                messages.error(request, 'Bank not found.')
            return redirect('admin_settings_banks')

        elif action == 'toggle_branch':
            branch_id = request.POST.get('branch_id')
            try:
                branch = BankBranch.objects.get(pk=branch_id)
                branch.active = not branch.active
                branch.save()
                messages.success(request, f"Branch {'activated' if branch.active else 'deactivated'}.")
            except BankBranch.DoesNotExist:
                messages.error(request, 'Branch not found.')
            return redirect('admin_settings_banks')

    banks = Bank.objects.prefetch_related('branches').all().order_by('name')

    context = {
        'nav': 'admin_settings_banks',
        'banks': banks,
        'bank_form': bank_form,
        'branch_form': branch_form,
        'edit_bank': edit_bank,
        'edit_branch': edit_branch,
    }
    return render(request, 'settings_banks.html', context)

@admin_check
def admin_settings_employers(request):
    employer_form = EmployerForm()
    edit_employer = None

    edit_employer_id = request.GET.get('edit_employer')
    if edit_employer_id:
        try:
            edit_employer = Employer.objects.get(pk=edit_employer_id)
            employer_form = EmployerForm(instance=edit_employer)
        except Employer.DoesNotExist:
            pass

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_employer':
            employer_form = EmployerForm(request.POST)
            if employer_form.is_valid():
                employer_form.save()
                messages.success(request, 'Employer added successfully.')
                return redirect('admin_settings_employers')

        elif action == 'edit_employer':
            employer_id = request.POST.get('employer_id')
            try:
                edit_employer = Employer.objects.get(pk=employer_id)
                employer_form = EmployerForm(request.POST, instance=edit_employer)
                if employer_form.is_valid():
                    employer_form.save()
                    messages.success(request, 'Employer updated successfully.')
                    return redirect('admin_settings_employers')
            except Employer.DoesNotExist:
                messages.error(request, 'Employer not found.')

        elif action == 'toggle_employer':
            employer_id = request.POST.get('employer_id')
            try:
                employer = Employer.objects.get(pk=employer_id)
                employer.active = not employer.active
                employer.save()
                messages.success(request, f"Employer {'activated' if employer.active else 'deactivated'}.")
            except Employer.DoesNotExist:
                messages.error(request, 'Employer not found.')
            return redirect('admin_settings_employers')

    employers = Employer.objects.all().order_by('name')

    context = {
        'nav': 'admin_settings_employers',
        'employers': employers,
        'employer_form': employer_form,
        'edit_employer': edit_employer,
    }
    return render(request, 'settings_employers.html', context)
