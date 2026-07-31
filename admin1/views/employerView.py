import datetime
import re
import requests
from decimal import Decimal
from weakref import ref
from io import BytesIO
from django.conf import settings
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render
from pyparsing import empty
from socket import gaierror
from accounts.models import User, UserProfile, StaffProfile, SMEProfile
from loan.models import Loan, LoanFile, Statement, Payment, PaymentUploads
from loan.forms import PaymentForm
from admin1.forms import AdminSettingsForm

from django.contrib.sites.shortcuts import get_current_site

#EMAIL SETTINGS
from django.template.loader import render_to_string
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.utils.html import strip_tags
#admin sender email
from admin1.models import AdminSettings, Location, Employer
sender = settings.DEFAULT_SENDER_EMAIL

#Class Based Views
from django.views.generic.base import View
from wkhtmltopdf.views import PDFTemplateResponse

from django.db.models import Sum, Q

from django.conf import settings
from accounts.functions import admin_check
from dcc.functions import check_client_in_dcc, get_loans_for_client, get_transactions_for_client
domain = settings.DOMAIN

from dcc.serializers import LoanSerializer, StatementSerializer

from collections import defaultdict

################ 
# START OF CODE
################

@admin_check
def employer_overview(request):

    user_profiles = UserProfile.objects.all()
    employers_dict = defaultdict(list)
    for profile in user_profiles:
        if profile.employer:
            employers_dict[profile.employer].append(profile)

    employers = {}
    for employer, profiles in employers_dict.items():
        total_loans = Loan.objects.filter(owner__in=profiles).count()
        repayment_total = Loan.objects.filter(owner__in=profiles).aggregate(Sum('repayment_amount'))['repayment_amount__sum'] or 0
        outstanding_total = Loan.objects.filter(owner__in=profiles).aggregate(Sum('total_outstanding'))['total_outstanding__sum'] or 0
        default_total = Loan.objects.filter(owner__in=profiles, status='DEFAULTED').count()
        arrears_total = Loan.objects.filter(owner__in=profiles).aggregate(Sum('total_arrears'))['total_arrears__sum'] or 0

        employers[employer] = {
            'profiles': profiles,
            'number_of_loans': total_loans,
            'repayment_total': repayment_total,
            'outstanding_total': outstanding_total,
            'default_total': default_total,
            'arrears_total': arrears_total
        }

    context = {
        'nav': 'employer_overview',
        'employers': employers,
        }
    
    return render(request, 'employer_overview.html', context)

def _loans_by_employer_data():
    """Build the Loans-per-Employer structure with per-loan detail and
    per-department (employer) totals. Shared by the page and its exports."""
    user_profiles = UserProfile.objects.all()
    employers_dict = defaultdict(list)
    for profile in user_profiles:
        if profile.employer:
            employers_dict[profile.employer].append(profile)

    employers = {}
    for employer, profiles in sorted(employers_dict.items()):
        loans = Loan.objects.filter(owner__in=profiles).exclude(funded_category='COMPLETED')
        loan_details = []
        totals = {
            'loan_amount': Decimal('0'), 'repayment_amount': Decimal('0'),
            'total_paid': Decimal('0'), 'total_outstanding': Decimal('0'),
            'number_of_defaults': 0, 'total_arrears': Decimal('0'),
        }
        for loan in loans:
            loan_details.append({
                'loan_ref': loan.ref,
                'first_name': loan.owner.first_name,
                'last_name': loan.owner.last_name,
                'status': loan.status,
                'loan_amount': loan.amount,
                'number_of_fortnights': loan.number_of_fortnights,
                'repayment_amount': loan.repayment_amount,
                'total_paid': loan.total_paid,
                'total_outstanding': loan.total_outstanding,
                'number_of_defaults': loan.number_of_defaults,
                'total_arrears': loan.total_arrears,
                'user_id': loan.owner.id,
            })
            totals['loan_amount'] += loan.amount or 0
            totals['repayment_amount'] += loan.repayment_amount or 0
            totals['total_paid'] += loan.total_paid or 0
            totals['total_outstanding'] += loan.total_outstanding or 0
            totals['number_of_defaults'] += loan.number_of_defaults or 0
            totals['total_arrears'] += loan.total_arrears or 0

        if loan_details:
            employers[employer] = {
                'profiles': profiles,
                'loans': loan_details,
                'totals': totals,
            }
    return employers


def _lbe_grand_total(employers):
    """Sum every employer's subtotal into one grand total across all employers."""
    gt = {'loan_count': 0, 'loan_amount': Decimal('0'), 'repayment_amount': Decimal('0'),
          'total_paid': Decimal('0'), 'total_outstanding': Decimal('0'),
          'number_of_defaults': 0, 'total_arrears': Decimal('0')}
    for data in employers.values():
        t = data['totals']
        gt['loan_count'] += len(data['loans'])
        gt['loan_amount'] += t['loan_amount']
        gt['repayment_amount'] += t['repayment_amount']
        gt['total_paid'] += t['total_paid']
        gt['total_outstanding'] += t['total_outstanding']
        gt['number_of_defaults'] += t['number_of_defaults']
        gt['total_arrears'] += t['total_arrears']
    return gt


@admin_check
def loans_by_employer(request):
    employers = _loans_by_employer_data()
    context = {
        'nav': 'loans_by_employer',
        'employers': employers,
        'grand_total': _lbe_grand_total(employers),
        'employer_count': len(employers),
        'domain': settings.DOMAIN,
    }
    return render(request, 'loans_by_employer.html', context)


_LBE_HEADERS = ['Loan Ref', 'Full Name', 'Status', 'Amount', 'No.of FNs',
                'Repayment', 'Total Paid', 'Outstanding', 'Defaults', 'Arrears']


def _lbe_flat_rows():
    """Flatten the report to (employer, rows, totals_row) blocks for export."""
    employers_data = _loans_by_employer_data()
    blocks = []
    for employer, data in employers_data.items():
        rows = []
        for l in data['loans']:
            rows.append([
                l['loan_ref'], f"{l['first_name']} {l['last_name']}", l['status'],
                l['loan_amount'], l['number_of_fortnights'], l['repayment_amount'],
                l['total_paid'], l['total_outstanding'], l['number_of_defaults'],
                l['total_arrears'],
            ])
        t = data['totals']
        totals_row = ['TOTAL', '', '', t['loan_amount'], '', t['repayment_amount'],
                      t['total_paid'], t['total_outstanding'], t['number_of_defaults'],
                      t['total_arrears']]
        blocks.append((employer, rows, totals_row))
    gt = _lbe_grand_total(employers_data)
    grand_row = [f"GRAND TOTAL ({gt['loan_count']} loans / {len(employers_data)} employers)", '', '',
                 gt['loan_amount'], '', gt['repayment_amount'], gt['total_paid'],
                 gt['total_outstanding'], gt['number_of_defaults'], gt['total_arrears']]
    return blocks, grand_row


@admin_check
def loans_by_employer_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Loans per Employer'
    bold = Font(bold=True)
    hdr_fill = PatternFill(fill_type='solid', start_color='344767', end_color='344767')
    hdr_font = Font(bold=True, color='FFFFFF')
    blocks, grand_row = _lbe_flat_rows()
    for employer, rows, totals_row in blocks:
        ws.append([employer])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=13)
        ws.append(_LBE_HEADERS)
        for c in ws[ws.max_row]:
            c.font = hdr_font; c.fill = hdr_fill
        for r in rows:
            ws.append([str(x) if not isinstance(x, (int, float, Decimal)) else float(x) for x in r])
        ws.append([str(x) if not isinstance(x, (int, float, Decimal)) else float(x) for x in totals_row])
        for c in ws[ws.max_row]:
            c.font = bold
        ws.append([])
    # Grand total across all employers.
    ws.append([str(x) if not isinstance(x, (int, float, Decimal)) else float(x) for x in grand_row])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, size=12); c.fill = PatternFill(fill_type='solid', start_color='E8ECF3', end_color='E8ECF3')
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="loans_per_employer_{datetime.date.today()}.xlsx"'
    return resp


@admin_check
def loans_by_employer_txt(request):
    import csv
    resp = HttpResponse(content_type='text/plain')
    resp['Content-Disposition'] = f'attachment; filename="loans_per_employer_{datetime.date.today()}.txt"'
    writer = csv.writer(resp, delimiter='\t')
    blocks, grand_row = _lbe_flat_rows()
    for employer, rows, totals_row in blocks:
        writer.writerow([employer])
        writer.writerow(_LBE_HEADERS)
        for r in rows:
            writer.writerow([str(x) for x in r])
        writer.writerow([str(x) for x in totals_row])
        writer.writerow([])
    writer.writerow([str(x) for x in grand_row])
    return resp


@admin_check
def loans_by_employer_pdf(request):
    filename = f'loans_per_employer_{datetime.date.today()}.pdf'
    block_list, grand_row = _lbe_flat_rows()
    blocks = [{'employer': e, 'rows': rows, 'totals': totals} for e, rows, totals in block_list]
    return PDFTemplateResponse(
        request=request,
        template='loans_by_employer_pdf.html',
        filename=filename,
        context={'headers': _LBE_HEADERS, 'blocks': blocks, 'grand_row': grand_row,
                 'generated': datetime.date.today().strftime('%d/%m/%Y')},
        show_content_in_browser=False,
        cmd_options={'orientation': 'Landscape', 'page-size': 'A4',
                     'margin-top': 10, 'margin-bottom': 12, 'margin-left': 8, 'margin-right': 8,
                     'footer-center': 'Page [page] of [topage]', 'footer-font-size': 8},
    )


@admin_check
def payroll_contacts(request):
    """List all registered Employer records with payroll officer contact details."""
    employers = Employer.objects.all().order_by('name')

    # Attach client count per employer (clients whose employer name matches)
    from django.db.models import Count
    employer_client_counts = {}
    for emp in employers:
        employer_client_counts[emp.id] = UserProfile.objects.filter(employer=emp.name).count()

    employer_data = [
        {
            'employer': emp,
            'client_count': employer_client_counts.get(emp.id, 0),
        }
        for emp in employers
    ]

    return render(request, 'payroll_contacts.html', {
        'nav': 'payroll_contacts',
        'employer_data': employer_data,
    })


def _payroll_rows():
    """Return (headers, rows) for export."""
    employers = Employer.objects.all().order_by('name')
    headers = ['#', 'Company / Organisation', 'Sector', 'Organisation Type',
               'Payroll Officer', 'Phone', 'Email', 'Office Address', 'Client Count', 'Status']
    rows = []
    for i, emp in enumerate(employers, start=1):
        client_count = UserProfile.objects.filter(employer=emp.name).count()
        rows.append([
            i,
            emp.name,
            emp.sector or '',
            emp.organisation_type or '',
            emp.payroll_officer_name or '',
            str(emp.work_phone) if emp.work_phone else '',
            emp.work_email or '',
            emp.office_address or '',
            client_count,
            'Active' if emp.active else 'Inactive',
        ])
    return headers, rows


@admin_check
def payroll_contacts_excel(request):
    """Export Payroll Contacts as an Excel (.xlsx) file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    headers, rows = _payroll_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll Contacts'

    # Header row styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(fill_type='solid', start_color='344767', end_color='344767')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    even_fill = PatternFill(fill_type='solid', start_color='EEF2FF', end_color='EEF2FF')
    for idx, row in enumerate(rows, start=2):
        ws.append(row)
        fill = even_fill if idx % 2 == 0 else PatternFill()
        for cell in ws[idx]:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
            if fill.fill_type:
                cell.fill = fill

    # Auto-fit column widths (approximate)
    col_widths = [5, 35, 12, 18, 25, 16, 35, 35, 12, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'payroll_contacts_{datetime.date.today()}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@admin_check
def payroll_contacts_pdf(request):
    """Export Payroll Contacts as a PDF via wkhtmltopdf."""
    employers = Employer.objects.all().order_by('name')
    employer_data = []
    for emp in employers:
        employer_data.append({
            'employer': emp,
            'client_count': UserProfile.objects.filter(employer=emp.name).count(),
        })

    filename = f'payroll_contacts_{datetime.date.today()}.pdf'
    return PDFTemplateResponse(
        request=request,
        template='payroll_contacts_pdf.html',
        filename=filename,
        context={
            'employer_data': employer_data,
            'generated': datetime.date.today().strftime('%d/%m/%Y'),
            'domain': settings.DOMAIN,
        },
        show_content_in_browser=False,
        cmd_options={
            'margin-top': 12,
            'margin-bottom': 12,
            'margin-left': 10,
            'margin-right': 10,
            'orientation': 'Landscape',
            'page-size': 'A4',
            'zoom': 1,
            'footer-center': 'Page [page] of [topage]',
            'footer-font-size': 8,
        },
    )